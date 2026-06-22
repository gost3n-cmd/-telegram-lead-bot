# -*- coding: utf-8 -*-
"""
CONTENT OPERATING SYSTEM
AI / Productivity / Digital Tools — Instagram Reels growth engine.

Architecture: 5 layers (Strategy, Psychology, Content Engine, Execution,
KPI & Analytics) + Content Bank deliverable + Reality Check appendix.
Run: python3 build_system.py  (writes ai_content_os.xlsx next to this file)
"""
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════════
# SHEET OWNERSHIP — защита данных бота (раздел 4 архитектуры)
#
# Excel-файл — общее хранилище двух независимых систем:
#   - System A: этот генератор — владеет 11 контентными листами
#   - System B: Telegram-бот — владеет листом Leads
#
# Правило (раздел 4.2): OWNED_SHEETS — единственный источник истины
# о том, какие листы вправе создавать, пересоздавать или удалять
# генератор. Любой лист вне этого списка для генератора не существует:
# не читается, не перебирается, не изменяется и не удаляется.
#
# Алгоритм (раздел 4.3): read-if-exists → rebuild owned → preserve foreign → save
# ════════════════════════════════════════════════════════════════════
_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _project_root)
from bot.config import BotConfig
_env_path = os.path.join(_project_root, ".env")
_config = BotConfig(_env_file=_env_path)
OUT_PATH = os.path.join(_project_root, _config.excel_file_path)

# Позитивный список владения (раздел 4.2) — 11 точных имён.
# Порядок = канонический порядок вкладок после reorder_owned_first.
OWNED_SHEETS = [
    "🏠 Home",
    "🎯 01 Strategy",
    "📱 02 Accounts",
    "🧠 03 Psychology",
    "🎬 04 Content Engine",
    "⚙️ 05 Pipeline",
    "👥 06 Team",
    "🛠️ 07 AI Stack",
    "📈 08 KPI Dashboard",
    "📚 09 Content Bank",
    "🔬 10 Reality Check",
]

if not os.path.exists(OUT_PATH):
    # Шаг 4.3.1: файла нет → создаём с нуля (поведение как раньше).
    # Убедиться, что родительский каталог существует.
    parent_dir = os.path.dirname(OUT_PATH)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)
    wb = Workbook()
    # У свежей Workbook() уже есть один пустой лист "Sheet" — он станет
    # ws0 (🏠 Home) ниже через wb.active. Ничего дополнительно делать не нужно.
else:
    # Шаг 4.3.2: файл существует → load_workbook (со всеми листами, включая чужие),
    # затем удалить ТОЛЬКО owned-листы перед их пересозданием.
    wb = load_workbook(OUT_PATH)
    for _name in OWNED_SHEETS:
        if _name in wb.sheetnames:
            wb.remove(wb[_name])
    # Ниже ws0=wb.active / wb.create_sheet(...) построят 11 owned-листов заново
    # неизменной логикой. Чужие листы (например, Leads) физически не затрагиваются.
    #
    # ВАЖНО: исходный код использует `ws0 = wb.active` и переназывает этот лист
    # в "🏠 Home". При load_workbook() без owned-листов wb.active указывал бы на
    # ЧУЖОЙ лист (Leads) — переназвание уничтожило бы его. Поэтому создаём один
    # пустой лист и делаем его активным: он и станет ws0 (🏠 Home), точно как при
    # свежей Workbook(). Это сохраняет неизменной всю логику ниже построчно.
    _seed = wb.create_sheet("_owned_seed_home_")
    wb.active = wb.sheetnames.index(_seed.title)




# ════════════════════════════════════════════════════════════════════
# DESIGN SYSTEM — one font, one neutral ink, one accent, status colors
# used ONLY where they encode meaning (automation level, virality tier,
# funnel stage, principle legend). No decorative rainbow fills.
# ════════════════════════════════════════════════════════════════════
FONT = "Calibri"

INK     = "0F172A"   # primary dark — hero titles
SLATE   = "1E293B"   # section headers / table headers
LINE    = "D9DEE6"   # borders
ROW_ALT = "F4F6F9"   # zebra stripe
WHITE   = "FFFFFF"
TEXT    = "1E293B"
MUTED   = "64748B"

ACCENT    = "2453E0"  # the one accent — emphasis, computed numbers, key facts
ACCENT_BG = "EAF0FE"

GOOD, GOOD_BG = "147D4B", "E7F6EE"   # status: positive / high / automated
WARN, WARN_BG = "B5650A", "FDF1E2"   # status: medium / caution
BAD,  BAD_BG  = "B3261E", "FBEAE9"   # status: low / risk / manual

# Layer identity — used only as sheet-tab color + thin kicker bar,
# never as cell-fill decoration inside tables.
L_STRAT = "2453E0"
L_PSY   = "6B3FA0"
L_ENG   = "0E7C7B"
L_EXEC  = "B5650A"
L_KPI   = "147D4B"
L_REAL  = "B3261E"

# ── low-level helpers ──────────────────────────────────────────────
def bd(color=LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cl(ws, row, col, val=None, bold=False, italic=False, sz=10, fc=TEXT,
       bg=None, align="left", valign="center", wrap=True, border=True,
       nf=None, indent=0):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.font = Font(bold=bold, italic=italic, size=sz, color=fc, name=FONT)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap, indent=indent)
    if border:
        c.border = bd()
    if nf:
        c.number_format = nf
    return c

def mspan(ws, r, c1, c2, val=None, **kw):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    return cl(ws, r, c1, val, **kw)

def hd(ws, row, col, text, bg=SLATE, fg=WHITE, sz=9.5, align="left"):
    return cl(ws, row, col, text, bold=True, sz=sz, fc=fg, bg=bg, align=align, border=True)

def wid(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

def sp(ws, row, h=8):
    ws.row_dimensions[row].height = h

def header_block(ws, c1, c2, kicker, title, subtitle, accent):
    """Three-row hero: kicker bar / dark title / muted subtitle. Returns next free row."""
    mspan(ws, 1, c1, c2, kicker, bold=True, sz=9, fc=WHITE, bg=accent,
          align="left", border=False, indent=1)
    rh(ws, 1, 20)
    mspan(ws, 2, c1, c2, title, bold=True, sz=16, fc=WHITE, bg=INK,
          align="left", border=False, indent=1)
    rh(ws, 2, 36)
    mspan(ws, 3, c1, c2, subtitle, italic=True, sz=10, fc=MUTED, bg=WHITE,
          align="left", border=False, indent=1, wrap=True)
    rh(ws, 3, 30)
    sp(ws, 4, 8)
    return 5

def section(ws, row, c1, c2, label, accent=SLATE, h=22):
    """Section divider bar. Returns next free row."""
    mspan(ws, row, c1, c2, label, bold=True, sz=10.5, fc=WHITE, bg=accent,
          align="left", border=False, indent=1)
    rh(ws, row, h)
    return row + 1

def note(ws, row, c1, c2, text, h=18, bg=WHITE, fc=MUTED):
    mspan(ws, row, c1, c2, text, italic=True, sz=8.5, fc=fc, bg=bg, border=False, indent=1)
    rh(ws, row, h)
    return row + 1

def freeze_below_header(ws, row):
    ws.freeze_panes = ws.cell(row=row, column=1)

# ════════════════════════════════════════════════════════════════════
# SHEET 0 — HOME / SYSTEM MAP
# ════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "🏠 Home"
ws0.sheet_view.showGridLines = False
ws0.sheet_properties.tabColor = INK
wid(ws0, [3, 24, 50, 32])

r = header_block(
    ws0, 1, 4,
    "CONTENT OPERATING SYSTEM",
    "🧠  AI / PRODUCTIVITY / DIGITAL TOOLS — REELS GROWTH OS",
    "Производственная система коротких видео для Instagram, построенная на "
    "6 психологических принципах удержания и распространения, а не на догадках.",
    INK,
)

r = section(ws0, r, 1, 4, "  СЕВЕРНАЯ ЗВЕЗДА", INK)
ns_items = [
    ("Цель", "100 000 000 просмотров / месяц"),
    ("Объём", "100 видео / день · 3 000 / месяц"),
    ("Экосистема", "15 нишевых Instagram-аккаунтов"),
    ("Логика модели", "Каждый сценарий = 1 из 6 психологических триггеров"),
]
for i, (k, v) in enumerate(ns_items):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws0, rr, 1, "", bg=bg, border=False)
    mspan(ws0, rr, 2, 2, k, bold=True, bg=ACCENT_BG, fc=ACCENT, align="left")
    mspan(ws0, rr, 3, 4, v, bg=bg, fc=TEXT)
    rh(ws0, rr, 20)
r += len(ns_items) + 1

r = section(ws0, r, 1, 4, "  АРХИТЕКТУРА: 5 СЛОЁВ + 2 ПРИЛОЖЕНИЯ", INK)
hd(ws0, r, 1, ""); hd(ws0, r, 2, "Слой"); hd(ws0, r, 3, "Назначение"); hd(ws0, r, 4, "Листы")
rh(ws0, r, 22)
r += 1

layers = [
    (L_STRAT, "L1 · STRATEGY", "Куда мы идём и насколько масштабируемся", "🎯 01 Strategy  ·  📱 02 Accounts"),
    (L_PSY,   "L2 · PSYCHOLOGY", "Почему человек останавливается, досматривает и делится", "🧠 03 Psychology"),
    (L_ENG,   "L3 · CONTENT ENGINE", "Как психология превращается в готовый сценарий", "🎬 04 Content Engine  ·  📚 09 Content Bank"),
    (L_EXEC,  "L4 · EXECUTION", "Как 1 сценарий становится 100 опубликованными видео в день", "⚙️ 05 Pipeline  ·  👥 06 Team  ·  🛠️ 07 AI Stack"),
    (L_KPI,   "L5 · KPI & ANALYTICS", "Как мы понимаем, что система действительно работает", "📈 08 KPI Dashboard"),
    (L_REAL,  "ПРИЛОЖЕНИЕ · REALITY CHECK", "Что может сломать систему в Instagram и как это компенсировать", "🔬 10 Reality Check"),
]
for i, (color, name, purpose, sheets) in enumerate(layers):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws0, rr, 1, "", bg=color, border=False)
    cl(ws0, rr, 2, name, bold=True, bg=bg, fc=TEXT)
    cl(ws0, rr, 3, purpose, bg=bg, fc=TEXT, sz=9.5)
    cl(ws0, rr, 4, sheets, bg=bg, fc=MUTED, sz=9, italic=True)
    rh(ws0, rr, 30)
r += len(layers) + 1

r = section(ws0, r, 1, 4, "  КАК ПОЛЬЗОВАТЬСЯ ФАЙЛОМ", SLATE)
howto = [
    "Двигайтесь по вкладкам слева направо — это и есть последовательность производства: Strategy → Psychology → Content Engine → Execution → KPI.",
    "Каждое видео в «09 Content Bank» явно ссылается на 1 из 6 принципов из «03 Psychology» — это не два независимых листа, а вход и выход одной модели.",
    "Цветной квадрат слева от названия слоя выше = цвет вкладки этого слоя внизу экрана — так проще ориентироваться, не открывая каждый лист.",
    "Перед публикацией контента, произведённого по системе, прочитайте «10 Reality Check» — он объясняет, что может занижать охват и как это обойти.",
]
for i, t in enumerate(howto):
    r = note(ws0, r, 1, 4, "•  " + t, h=28, bg=ROW_ALT if i % 2 == 0 else WHITE, fc=TEXT)

freeze_below_header(ws0, 5)

print("Sheet 0 (Home) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — STRATEGY (Layer 1)
# ════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("🎯 01 Strategy")
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = L_STRAT
wid(ws1, [30, 16, 44])

r = header_block(
    ws1, 1, 3,
    "LAYER 1 · STRATEGY",
    "🎯  ЦЕЛИ, KPI-КАСКАД И ЛОГИКА МАСШТАБИРОВАНИЯ",
    "Куда движется система и какими проверяемыми шагами она туда доходит.",
    L_STRAT,
)

# --- KPI cascade with real formulas (inputs in blue-accent, outputs computed) ---
r = section(ws1, r, 1, 3, "  ВХОДНЫЕ ДОПУЩЕНИЯ", L_STRAT)
hd(ws1, r, 1, "Параметр"); hd(ws1, r, 2, "Значение"); hd(ws1, r, 3, "Комментарий")
rh(ws1, r, 22)
r += 1
in_start = r
inputs = [
    ("Целевые просмотры / месяц", 100000000, "Северная метрика всей системы", '#,##0'),
    ("Видео в день", 100, "Производственная мощность на полной фазе", '#,##0'),
    ("Дней в расчётном месяце", 30, "Стандартный месяц для планирования", '#,##0'),
    ("Аккаунтов в экосистеме", 15, "См. 02 Accounts — нишевая сегментация", '#,##0'),
]
for i, (lbl, val, cmt, fmt) in enumerate(inputs):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws1, rr, 1, lbl, bg=bg, fc=TEXT)
    cl(ws1, rr, 2, val, bg=ACCENT_BG, fc=ACCENT, bold=True, align="center", nf=fmt)
    cl(ws1, rr, 3, cmt, bg=bg, fc=MUTED, italic=True, sz=9)
    rh(ws1, rr, 20)
TARGET_VIEWS_ROW, VPD_ROW, DAYS_ROW, ACC_ROW = in_start, in_start + 1, in_start + 2, in_start + 3
r += len(inputs) + 1

r = section(ws1, r, 1, 3, "  РАСЧЁТНЫЕ ПОКАЗАТЕЛИ (формулы)", SLATE)
hd(ws1, r, 1, "Показатель"); hd(ws1, r, 2, "Значение"); hd(ws1, r, 3, "Формула / логика")
rh(ws1, r, 22)
r += 1
out_start = r
outputs = [
    ("Видео в месяц", f"=B{VPD_ROW}*B{DAYS_ROW}", "Видео/день × Дни"),
    ("Видео на 1 аккаунт в день", f"=B{VPD_ROW}/B{ACC_ROW}", "Видео/день ÷ Аккаунтов"),
    ("Целевое среднее просмотров / видео", f"=B{TARGET_VIEWS_ROW}/B{out_start}", "Цель ÷ Видео/месяц"),
    ("Целевой охват на 1 аккаунт / месяц", f"=B{TARGET_VIEWS_ROW}/B{ACC_ROW}", "Цель ÷ Аккаунтов"),
]
for i, (lbl, formula, cmt) in enumerate(outputs):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws1, rr, 1, lbl, bg=bg, fc=TEXT)
    cl(ws1, rr, 2, formula, bg=bg, fc=ACCENT, bold=True, align="center", nf='#,##0')
    cl(ws1, rr, 3, cmt, bg=bg, fc=MUTED, italic=True, sz=9)
    rh(ws1, rr, 20)
r += len(outputs) + 1

# --- Scaling roadmap ---
r = section(ws1, r, 1, 3, "  ROADMAP МАСШТАБИРОВАНИЯ КОНТЕНТА", L_STRAT)
hd(ws1, r, 1, "Фаза / срок"); hd(ws1, r, 2, "Видео/день"); hd(ws1, r, 3, "Фокус и критерий перехода дальше")
rh(ws1, r, 22)
r += 1
phases = [
    ("Фаза 0 · Инфраструктура · Нед. 1–2", "0", "Создать и прогреть 15 аккаунтов, собрать библиотеку хуков и шаблонов промптов. Переход: все инструменты подключены, 15 аккаунтов активны."),
    ("Фаза 1 · Тест · Нед. 3–4", "30", "5 аккаунтов, A/B хуков, поиск winning-форматов. Переход: ≥1 видео с органическим охватом >10× от подписчиков аккаунта."),
    ("Фаза 2 · Масштаб I · Мес. 2", "60", "Запуск всех 15 аккаунтов, полная автоматизация войсовера/субтитров. Переход: стабильные 60 видео/день 14 дней подряд без падения retention."),
    ("Фаза 3 · Полный масштаб · Мес. 3", "100", "Pipeline «сценарий → публикация» < 20 мин, запуск нативного промо-слота. Переход: 100 видео/день держится 21 день, очередь kill-листа работает еженедельно."),
    ("Фаза 4 · Оптимизация · Мес. 4–5", "100", "Удвоение топ-20% форматов, отключение слабых ниш, тест платного буста на лидерах. Переход: CPL устойчиво падает 2 месяца подряд."),
    ("Фаза 5 · Цель · Мес. 6", "100", "100 видео/день на максимальной автоматизации, полная интеграция воронки. Цель: 100M просмотров/мес, см. 08 KPI Dashboard."),
]
for i, (phase, vpd, desc) in enumerate(phases):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws1, rr, 1, phase, bold=True, bg=bg, fc=L_STRAT, sz=9.5)
    cl(ws1, rr, 2, vpd, bold=True, bg=bg, align="center")
    cl(ws1, rr, 3, desc, bg=bg, fc=TEXT, sz=9)
    rh(ws1, rr, 46)
r += len(phases) + 1

# --- Strategic pillars ---
r = section(ws1, r, 1, 3, "  5 СТРАТЕГИЧЕСКИХ ПРИНЦИПОВ", SLATE)
hd(ws1, r, 1, "Принцип"); hd(ws1, r, 2, ""); hd(ws1, r, 3, "Суть")
rh(ws1, r, 22)
r += 1
pillars = [
    ("1 · Volume × Automation", "AI-пайплайн производит объём, человек контролирует только QC-шлюз перед публикацией (см. 05 Pipeline)."),
    ("2 · Niche Segmentation", "15 аккаунтов = 15 чётких тем. Критично для алгоритмических topic-controls Instagram (см. 10 Reality Check)."),
    ("3 · Data-Driven Kill / Scale", "Еженедельно: нижние 20% по retention — в архив, верхние 20% — удваиваем формат (см. 08 KPI Dashboard)."),
    ("4 · Hook-First Production", "Сценарий не пишется, пока хук не закреплён за одним из 6 принципов (см. 03 Psychology, 04 Content Engine)."),
    ("5 · Funnel Architecture", "Холодный охват через психологию → тёплая аудитория через идентичность/регулярность → конверсия в промо-слоте."),
]
for i, (p, e) in enumerate(pillars):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws1, rr, 1, p, bold=True, bg=bg, fc=L_STRAT, sz=9.5)
    mspan(ws1, rr, 2, 3, e, bg=bg, fc=TEXT, sz=9)
    rh(ws1, rr, 30)

freeze_below_header(ws1, 5)
print("Sheet 1 (Strategy) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — ACCOUNT ARCHITECTURE (Layer 1)
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📱 02 Accounts")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = L_STRAT
wid(ws2, [4, 20, 24, 11, 28, 24, 16, 16])

r = header_block(
    ws2, 1, 8,
    "LAYER 1 · STRATEGY",
    "📱  АРХИТЕКТУРА АККАУНТОВ — 15 НИШЕВЫХ КАНАЛОВ",
    "Сегментация по нишам напрямую служит Mere Exposure (узнаваемость формата) "
    "и алгоритмическим topic-controls Instagram — см. 10 Reality Check.",
    L_STRAT,
)

r = section(ws2, r, 1, 8, "  ЭКОСИСТЕМА АККАУНТОВ", L_STRAT)
heads = ["#", "Handle", "Ниша", "Видео/день", "Контент-фокус", "Аудитория", "Этап воронки", "Цель подписч. (мес.6)"]
for ci, h in enumerate(heads, 1):
    hd(ws2, r, ci, h)
rh(ws2, r, 24)
r += 1
acc_start = r

FUNNEL = {
    "Холодный": (ACCENT_BG, ACCENT),
    "Тёплый": (WARN_BG, WARN),
    "Горячий": (GOOD_BG, GOOD),
    "Конверсия": (BAD_BG, BAD),
}
accounts = [
    (1, "@ai_tools_daily", "Обзоры AI-инструментов", 7, "Туториалы, демо, сравнения интерфейсов", "IT-специалисты, предприниматели", "Холодный", 500000),
    (2, "@chatgpt_hacks_", "Лайфхаки ChatGPT", 7, "Скрытые функции, быстрые трюки", "Широкая аудитория 18–35", "Холодный", 600000),
    (3, "@ai_news_fast", "AI-новости за 60 сек", 7, "Breaking news, релизы, рекорды", "Tech-энтузиасты, широкая", "Холодный", 450000),
    (4, "@midjourney_art_", "AI-арт Midjourney/Leonardo", 6, "Шоукейсы, процесс генерации", "Художники, дизайнеры", "Тёплый", 350000),
    (5, "@runway_visions", "AI видео-арт Runway/Pika", 6, "Сгенерированные клипы, wow-эффект", "Широкая, tech-аудитория", "Тёплый", 300000),
    (6, "@ai_for_biz_", "AI для бизнеса", 6, "Кейсы, ROI, бизнес-применения", "Предприниматели, менеджеры", "Горячий", 300000),
    (7, "@ai_vs_human_", "AI против человека", 7, "Эксперименты, прямые сравнения", "Широкая, развлекательная", "Холодный", 500000),
    (8, "@ai_money_tips", "Заработок с AI", 7, "Способы монетизации AI-навыков", "Фрилансеры, предприниматели", "Горячий", 400000),
    (9, "@ai_funny_fails", "Приколы и фейлы AI", 6, "Юмор, неожиданные результаты", "Широкая аудитория 16–40", "Холодный", 450000),
    (10, "@learn_ai_fast", "Обучение AI с нуля", 6, "Объяснения тем простым языком", "Студенты, новички", "Конверсия", 300000),
    (11, "@ai_productive_", "Продуктивность с AI", 7, "Рабочие лайфхаки, автоматизация", "Офисные сотрудники, HR", "Тёплый", 350000),
    (12, "@ai_code_tricks", "AI для разработчиков", 6, "Coding с AI-ассистентами", "Разработчики, IT-студенты", "Тёплый", 250000),
    (13, "@ai_marketing_", "AI в маркетинге", 6, "Кейсы, инструменты, ROI-расчёты", "Маркетологи, SMM-специалисты", "Горячий", 250000),
    (14, "@future_ai_now_", "Будущее AI: прогнозы", 6, "Тренды, предсказания, аналитика", "Tech-энтузиасты, инвесторы", "Тёплый", 300000),
    (15, "@[бренд_продукта]", "Промо-аккаунт продукта", 5, "Обучение, кейсы, отзывы клиентов", "Потенциальные клиенты 18–45", "Конверсия", 200000),
]
for i, (num, handle, niche, vpd, focus, aud, stage, subs) in enumerate(accounts):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws2, rr, 1, num, bold=True, bg=bg, align="center", fc=ACCENT)
    cl(ws2, rr, 2, handle, bold=True, bg=bg, fc=INK)
    cl(ws2, rr, 3, niche, bg=bg, fc=TEXT, sz=9)
    cl(ws2, rr, 4, vpd, bold=True, bg=bg, align="center")
    cl(ws2, rr, 5, focus, bg=bg, fc=TEXT, sz=9)
    cl(ws2, rr, 6, aud, bg=bg, fc=TEXT, sz=9)
    sbg, sfc = FUNNEL[stage]
    cl(ws2, rr, 7, stage, bg=sbg, fc=sfc, bold=True, align="center", sz=9)
    cl(ws2, rr, 8, subs, bold=True, bg=bg, fc=GOOD, align="center", nf='#,##0')
    rh(ws2, rr, 30)
acc_end = r + len(accounts) - 1
r = acc_end + 1

mspan(ws2, r, 1, 3, "ИТОГО ПО ЭКОСИСТЕМЕ", bold=True, fc=WHITE, bg=INK, align="right", border=False, indent=1)
cl(ws2, r, 4, f"=SUM(D{acc_start}:D{acc_end})", bold=True, bg=INK, fc=WHITE, align="center")
mspan(ws2, r, 5, 6, "видео/день суммарно по всем 15 аккаунтам", fc=WHITE, bg=INK, sz=8.5, italic=True, border=False, indent=1)
mspan(ws2, r, 7, 7, "Σ подписч.:", fc=WHITE, bg=INK, sz=8.5, align="right", border=False)
cl(ws2, r, 8, f"=SUM(H{acc_start}:H{acc_end})", bold=True, bg=INK, fc=WHITE, align="center", nf='#,##0')
rh(ws2, r, 24)
r += 2

r = note(ws2, r, 1, 8, "Легенда «Этап воронки»:  Холодный = первый контакт (психология/виральность)   ·   "
         "Тёплый = повторный контакт, узнаваемость (Mere Exposure)   ·   Горячий = готов к действию   ·   "
         "Конверсия = прямой призыв к продукту/школе.", h=24)

freeze_below_header(ws2, 5)
print("Sheet 2 (Accounts) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 3 — PSYCHOLOGY (Layer 2) — the core of the model
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🧠 03 Psychology")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = L_PSY
wid(ws3, [22, 30, 30, 28, 34])

r = header_block(
    ws3, 1, 5,
    "LAYER 2 · PSYCHOLOGY",
    "🧠  6 ПРИНЦИПОВ УДЕРЖАНИЯ И ВИРАЛЬНОСТИ",
    "Это ядро модели, а не дополнение: каждый сценарий в 09 Content Bank "
    "явно закодирован одним из этих шести принципов.",
    L_PSY,
)

PRINCIPLES = [
    ("1 · Mere Exposure Effect", "Узнаваемые форматы важнее оригинальности",
     "Чем чаще мозг встречает знакомый стимул, тем больше он ему доверяет — "
     "даже без новой информации.",
     "Единый визуальный шаблон (заставка, шрифт, голос, структура подачи) "
     "повторяется в каждом видео аккаунта/рубрики, чтобы зритель узнавал "
     "формат за доли секунды.",
     "Узнавание, чувство комфорта, снижение скепсиса — «я уже знаю, что это работает».",
     "Рубрика «AI за 60 секунд» открывается одной и той же заставкой и фразой "
     "ведущего каждый выпуск — зритель остаётся, потому что знает структуру."),
    ("2 · Curiosity Trap", "Информационный разрыв",
     "Когда человек осознаёт зазор между тем, что он знает, и тем, что хочет "
     "узнать, возникает дискомфорт, который закрывается только досмотром.",
     "Хук озвучивает открытую петлю («один инструмент изменил всё») и удерживает "
     "ответ до 70–90% хронометража видео.",
     "Напряжение, предвкушение, навязчивое желание узнать развязку.",
     "«Есть настройка ChatGPT, о которой не знают 99% — досмотри до конца» — "
     "разгадка даётся только на последних секундах."),
    ("3 · Identity Selling", "Продажа идентичности, не продукта",
     "Люди не покупают функции — они покупают версию себя, к которой стремятся "
     "или которой боятся не соответствовать.",
     "Контент формулируется как маркер принадлежности к группе («то, что уже "
     "знают AI-пользователи 1%»), а не как нейтральный обзор функции.",
     "Стремление к статусу, страх отстать, защита самооценки, племенная принадлежность.",
     "«Если ты ещё закрываешь отчёты вручную — это видео для тебя» — "
     "зритель смотрит, чтобы перестать быть «тем самым» отстающим."),
    ("4 · Authority / Credentials", "Доверие за 2 секунды",
     "Зритель решает, доверять ли источнику, в первые 1–2 секунды — до того, "
     "как успевает обработать содержание логически.",
     "Первый кадр показывает доказательство компетенции: экран интерфейса, "
     "конкретную цифру результата, а не лицо, рассказывающее со слов.",
     "Эвристика доверия, снижение скепсиса — «этот человек/аккаунт разбирается».",
     "«Я протестировал 47 AI-инструментов, чтобы вам не пришлось» с реальным "
     "скринкастом сравнения вместо разговора на камеру."),
    ("5 · Social Sharing", "Контент, которым хотят делиться",
     "Контент распространяется не потому, что он «хороший», а потому что "
     "пересылка выгодна самому отправителю — делает его полезным/смешным/в теме.",
     "В payoff закладывается момент, который хочется переслать конкретному "
     "человеку («перешли тому, кто до сих пор делает это руками»).",
     "Социальная валюта — «я выгляжу полезным/осведомлённым, если поделюсь этим».",
     "Видео-сравнение «ChatGPT vs Claude в одной задаче» с подписью "
     "«сохрани — пригодится в понедельник» рассчитано на сейвы и репосты в директ."),
    ("6 · Storytelling", "Hook → Problem → Story → Resolution",
     "Нарративная структура удерживает внимание и врезается в память сильнее, "
     "чем плоская подача фактов, потому что незавершённая история психологически "
     "некомфортна для прерывания.",
     "Даже 30-секундный обзор инструмента строится как мини-история: завязка "
     "(проблема) → конкретное действие героя → развязка с измеримым результатом.",
     "Нарративная вовлечённость, потребность в завершении, эмоциональная память.",
     "«Чуть не сорвал дедлайн (hook) → отчёт на 6 часов работы (problem) → нашёл "
     "это в 11 вечера (story) → закрыл за 12 минут, вот как именно (resolution)»."),
]

r = section(ws3, r, 1, 5, "  ПРИНЦИП → ПРИМЕНЕНИЕ → ТРИГГЕРЫ → ПРИМЕР", L_PSY)
heads = ["Принцип", "Суть", "Применение в Shorts", "Психологические триггеры", "Пример применения"]
for ci, h in enumerate(heads, 1):
    hd(ws3, r, ci, h)
rh(ws3, r, 24)
r += 1
for i, (name, essence, defn, app, trig, ex) in enumerate(PRINCIPLES):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws3, rr, 1, name, bold=True, bg=bg, fc=L_PSY, sz=9.5)
    cl(ws3, rr, 2, essence + ".\n" + defn, bg=bg, fc=TEXT, sz=8.5)
    cl(ws3, rr, 3, app, bg=bg, fc=TEXT, sz=8.5)
    cl(ws3, rr, 4, trig, bg=bg, fc=TEXT, sz=8.5)
    cl(ws3, rr, 5, ex, bg=bg, fc=MUTED, sz=8.5, italic=True)
    rh(ws3, rr, 88)
r += len(PRINCIPLES) + 1

# Impact matrix
r = section(ws3, r, 1, 5, "  ВЛИЯНИЕ НА RETENTION И VIRALITY", SLATE)
hd(ws3, r, 1, "Принцип"); hd(ws3, r, 2, "Влияние на Retention"); hd(ws3, r, 3, "Влияние на Virality")
mspan(ws3, r, 4, 5, "Когда использовать в первую очередь", bold=True, sz=9.5, fc=WHITE, bg=SLATE, align="left")
rh(ws3, r, 22)
r += 1
TIER = {"Очень высокое": (ACCENT_BG, ACCENT), "Высокое": (GOOD_BG, GOOD), "Среднее": (WARN_BG, WARN), "Низкое-среднее": (ROW_ALT, MUTED)}
impact = [
    ("Mere Exposure", "Высокое", "Низкое-среднее", "Постоянные рубрики, серийный контент, фирменные открывашки аккаунта."),
    ("Curiosity Trap", "Очень высокое", "Среднее", "Хук видео, особенно для туториалов и сравнений."),
    ("Identity Selling", "Среднее", "Высокое", "Контент для тёплой/горячей аудитории, нативный промо-слот."),
    ("Authority / Credentials", "Высокое", "Среднее", "Первый кадр любого видео — независимо от пилларa."),
    ("Social Sharing", "Низкое-среднее", "Очень высокое", "Listicle/сравнения и payoff-моменты, рассчитанные на пересылку."),
    ("Storytelling", "Очень высокое", "Высокое", "Базовый каркас сценария — применяется почти всегда поверх остальных."),
]
for i, (name, ret, vir, when) in enumerate(impact):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws3, rr, 1, name, bold=True, bg=bg, fc=L_PSY, sz=9.5)
    rbg, rfc = TIER[ret]
    cl(ws3, rr, 2, ret, bg=rbg, fc=rfc, bold=True, align="center", sz=9)
    vbg, vfc = TIER[vir]
    cl(ws3, rr, 3, vir, bg=vbg, fc=vfc, bold=True, align="center", sz=9)
    mspan(ws3, rr, 4, 5, when, bg=bg, fc=TEXT, sz=9)
    rh(ws3, rr, 24)

freeze_below_header(ws3, 5)
print("Sheet 3 (Psychology) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 4 — CONTENT ENGINE (Layer 3)
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🎬 04 Content Engine")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = L_ENG
wid(ws4, [16, 26, 20, 38])

r = header_block(
    ws4, 1, 4,
    "LAYER 3 · CONTENT ENGINE",
    "🎬  SHORT-FORM SYSTEM: АНАТОМИЯ, ФОРМУЛЫ, ХУКИ",
    "Как принципы из 03 Psychology становятся производимым сценарием за 3 минуты.",
    L_ENG,
)

# Anatomy
r = section(ws4, r, 1, 4, "  АНАТОМИЯ РОЛИКА 0–30 СЕК", L_ENG)
hd(ws4, r, 1, "Таймкод"); hd(ws4, r, 2, "Назначение блока"); hd(ws4, r, 3, "Ведущий принцип"); hd(ws4, r, 4, "Типичная ошибка")
rh(ws4, r, 22)
r += 1
anatomy = [
    ("0:00–0:02", "Hook — паттерн-прерывание", "Curiosity Trap, Authority", "Хук слишком общий, нет конкретной цифры/факта в первом кадре."),
    ("0:02–0:05", "Обещание / ставки", "Curiosity Trap", "Обещание шире, чем реальный payoff — убивает доверие к следующему видео."),
    ("0:05–0:22", "Problem → Story / демонстрация", "Storytelling, Identity Selling", "Демонстрация без сюжета — список фактов вместо истории с героем."),
    ("0:22–0:27", "Resolution — payoff", "Authority, Social Sharing", "Слабый, неконкретный вывод без измеримого результата."),
    ("0:27–0:30", "CTA / loop", "Social Sharing, Mere Exposure", "Общий CTA «подписывайся» вместо конкретного действия (сохрани/перешли/пришли)."),
]
for i, (t, purpose, princ, mistake) in enumerate(anatomy):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws4, rr, 1, t, bold=True, bg=bg, align="center", fc=L_ENG)
    cl(ws4, rr, 2, purpose, bg=bg, fc=TEXT, sz=9)
    cl(ws4, rr, 3, princ, bg=bg, fc=ACCENT, sz=9, bold=True)
    cl(ws4, rr, 4, mistake, bg=bg, fc=MUTED, sz=8.5, italic=True)
    rh(ws4, rr, 30)
r += len(anatomy) + 1

# Hook formula library
r = section(ws4, r, 1, 4, "  БИБЛИОТЕКА ХУКОВ — 15 ФОРМУЛ", SLATE)
hd(ws4, r, 1, "#"); hd(ws4, r, 2, "Формула хука"); hd(ws4, r, 3, "Принцип"); hd(ws4, r, 4, "Пример заполнения")
rh(ws4, r, 22)
r += 1
hooks = [
    ("Я протестировал [N] [категория], и только один [результат]", "Authority + Curiosity", "Я протестировал 12 AI-планировщиков, и только один реально освободил мне вечер"),
    ("[N]% людей не знают, что [инструмент] умеет [X]", "Curiosity Trap", "92% людей не знают, что в Notion AI есть голосовой ввод задач"),
    ("Не [действие], пока не увидишь это", "Curiosity Trap", "Не плати за Canva Pro, пока не увидишь это"),
    ("Если ты [идентичность] — это изменит твою неделю", "Identity Selling", "Если ты фрилансер на удалёнке — это изменит твою неделю"),
    ("POV: ты [ситуация], а потом находишь [инструмент]", "Storytelling", "POV: дедлайн через час, а потом ты находишь это"),
    ("[N] секунд — и [результат], без [боль]", "Curiosity + Authority", "47 секунд — и готовый отчёт, без единой формулы Excel"),
    ("Это видео для тех, кто ещё [боль]", "Identity Selling", "Это видео для тех, кто ещё вручную переносит задачи из почты"),
    ("Удалил [N] приложений после того, как нашёл это", "Authority + Curiosity", "Удалил 4 приложения после того, как нашёл это"),
    ("Сравнил [A] и [B] — разница в [метрика] шокировала", "Social Sharing", "Сравнил ChatGPT и Claude в одном письме — разница в тоне шокировала"),
    ("То, что [инсайдеры] не расскажут тебе бесплатно", "Authority", "То, что AI-агентства не расскажут тебе бесплатно"),
    ("Я [боль], пока не сделал одно действие", "Storytelling", "Я терял 2 часа в день на отчёты, пока не сделал одно действие"),
    ("Сохрани это — пригодится [событие]", "Social Sharing", "Сохрани это — пригодится в понедельник утром"),
    ("[Инструмент] только что выпустил [фичу], и это меняет всё", "Curiosity + Authority", "ChatGPT только что выпустил функцию, и это меняет всё"),
    ("Перешли тому, кто до сих пор [боль]", "Social Sharing", "Перешли тому, кто до сих пор печатает письма с нуля"),
    ("Рубрика «[название]» #[N] — сегодня тестируем [тема]", "Mere Exposure", "Рубрика «AI за 60 секунд» #18 — сегодня тестируем Gamma"),
]
for i, (formula, princ, ex) in enumerate(hooks):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws4, rr, 1, i + 1, bold=True, bg=bg, align="center", fc=ACCENT)
    cl(ws4, rr, 2, formula, bg=bg, fc=TEXT, sz=9)
    cl(ws4, rr, 3, princ, bg=bg, fc=L_PSY, sz=8.5, bold=True)
    cl(ws4, rr, 4, ex, bg=bg, fc=MUTED, sz=8.5, italic=True)
    rh(ws4, rr, 26)
r += len(hooks) + 1

# Script formulas by content type
r = section(ws4, r, 1, 4, "  ФОРМУЛЫ СЦЕНАРИЕВ ПО ТИПУ КОНТЕНТА", L_ENG)
hd(ws4, r, 1, "Тип"); hd(ws4, r, 2, "Hook"); hd(ws4, r, 3, "Problem → Story"); hd(ws4, r, 4, "Resolution")
rh(ws4, r, 22)
r += 1
script_formulas = [
    ("Туториал", "«[N] секунд — и [результат]»", "Зритель не знает шага → показываем экран в реальном времени, без монтажных склеек на ключевом действии", "Финальный результат на экране + 1 фраза «вот и всё»"),
    ("Сравнение", "«[A] vs [B] — разница в [метрика]»", "Одна и та же задача даётся обоим инструментам параллельно", "Чёткий победитель по 1 измеримому критерию, не по «вкусу»"),
    ("Новость", "«[Инструмент] только что [событие]»", "Контекст за 3 секунды: что было раньше / что изменилось", "Что это значит для зрителя лично, прямо сейчас"),
    ("Шоукейс", "«Посмотри, что умеет [инструмент]»", "Нарастание: от простого результата к неожиданному", "Самый сильный результат — последним кадром, без слов"),
    ("История/признание", "«Я [боль], пока не [действие]»", "Конкретная ситуация с провалом → момент находки инструмента", "Измеримый итог + что изменилось в привычке"),
]
for i, (t, h_, ps, res) in enumerate(script_formulas):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws4, rr, 1, t, bold=True, bg=bg, fc=L_ENG, sz=9.5)
    cl(ws4, rr, 2, h_, bg=bg, fc=TEXT, sz=8.5)
    cl(ws4, rr, 3, ps, bg=bg, fc=TEXT, sz=8.5)
    cl(ws4, rr, 4, res, bg=bg, fc=TEXT, sz=8.5)
    rh(ws4, rr, 40)
r += len(script_formulas) + 1

# Content pillars
r = section(ws4, r, 1, 4, "  КОНТЕНТ-ПИЛЛАРЫ И РАСПРЕДЕЛЕНИЕ (на 1 аккаунт · 7 видео/день)", SLATE)
hd(ws4, r, 1, "Тип контента"); hd(ws4, r, 2, "Доля"); hd(ws4, r, 3, "Ведущий принцип"); hd(ws4, r, 4, "Вирал-потенциал")
rh(ws4, r, 22)
r += 1
VTIER = {"Максимальный": (ACCENT_BG, ACCENT), "Очень высокий": (GOOD_BG, GOOD), "Высокий": (WARN_BG, WARN), "Средний": (ROW_ALT, MUTED)}
pillars_dist = [
    ("Туториал / How-To", 0.25, "Authority + Curiosity Trap", "Высокий"),
    ("AI-новости", 0.20, "Curiosity Trap + Mere Exposure", "Очень высокий"),
    ("Сравнения", 0.15, "Social Sharing", "Очень высокий"),
    ("Шоукейс / Wow-контент", 0.15, "Authority + Storytelling", "Очень высокий"),
    ("Продуктивность / лайфхаки", 0.10, "Identity Selling", "Высокий"),
    ("Юмор / AI-фейлы", 0.10, "Social Sharing", "Максимальный"),
    ("Нативный промо", 0.05, "Identity Selling", "Средний"),
]
for i, (t, pct, princ, vir) in enumerate(pillars_dist):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws4, rr, 1, t, bold=True, bg=bg, fc=TEXT, sz=9)
    cl(ws4, rr, 2, pct, bg=bg, fc=ACCENT, bold=True, align="center", nf='0%')
    cl(ws4, rr, 3, princ, bg=bg, fc=L_PSY, sz=8.5, bold=True)
    vbg, vfc = VTIER[vir]
    cl(ws4, rr, 4, vir, bg=vbg, fc=vfc, bold=True, align="center", sz=9)
    rh(ws4, rr, 24)

freeze_below_header(ws4, 5)
print("Sheet 4 (Content Engine) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 5 — EXECUTION: PIPELINE (Layer 4)
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⚙️ 05 Pipeline")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = L_EXEC
wid(ws5, [4, 22, 30, 22, 13, 13, 14, 22])

r = header_block(
    ws5, 1, 8,
    "LAYER 4 · EXECUTION",
    "⚙️  ПРОИЗВОДСТВЕННЫЙ ПАЙПЛАЙН — END TO END",
    "Путь одного сценария от идеи до публикации на 100 видео/день.",
    L_EXEC,
)

r = section(ws5, r, 1, 8, "  8 ШАГОВ ПРОИЗВОДСТВА", L_EXEC)
heads = ["#", "Этап", "AI-инструменты", "Ответственный", "Мин/видео", "Часов/день", "Автоматизация", "Результат"]
for ci, h in enumerate(heads, 1):
    hd(ws5, r, ci, h)
rh(ws5, r, 24)
r += 1
AUTO = {"Полная": (GOOD_BG, GOOD), "Высокая": (GOOD_BG, GOOD), "Средняя": (WARN_BG, WARN), "Частичная": (WARN_BG, WARN), "Ручная": (BAD_BG, BAD)}
pipe = [
    (1, "Тренд-ресёрч", "Google Trends, TrendTok, Creative Center", "Тренд-ресёрчер", "1,2", "4", "Частичная", "100 утверждённых тем дня"),
    (2, "Генерация скрипта", "Claude / GPT-4o + шаблоны промптов", "Prompt-инженер", "3", "5 (batch)", "Высокая", "100 скриптов 30–60 сек"),
    (3, "Хук + обложка", "GPT-4o + база A/B хуков", "Контент-стратег", "1,5", "2,5", "Высокая", "100 хуков + темы обложек"),
    (4, "Войсовер", "ElevenLabs (batch API)", "Авто", "2", "авто", "Полная", "100 аудиодорожек"),
    (5, "Видеопродакшн", "CapCut auto-cut + HeyGen/Runway", "Видеоредактор", "8", "13", "Средняя", "100 черновых видео"),
    (6, "Субтитры + оформление", "CapCut auto-subs + Canva-шаблоны", "Видеоредактор", "2", "3,5", "Высокая", "100 готовых Reels"),
    (7, "QC-проверка", "Ручной spot-check 20%", "Контент-директор + стратег", "0,5", "1", "Ручная", "<10% уходит на переделку"),
    (8, "Публикация", "Later / Buffer / Meta Business Suite", "Менеджер аккаунтов", "0,5", "1", "Высокая", "100 запланированных постов"),
]
for i, row_data in enumerate(pipe):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    for ci, val in enumerate(row_data, 1):
        kw = dict(bg=bg, sz=8.8)
        if ci == 1:
            kw.update(bold=True, align="center", fc=ACCENT)
        elif ci in (5, 6):
            kw.update(align="center")
        elif ci == 7:
            abg, afc = AUTO.get(str(val), (bg, TEXT))
            kw.update(bg=abg, fc=afc, bold=True, align="center")
        cl(ws5, rr, ci, val, **kw)
    rh(ws5, rr, 36)
r += len(pipe)

mspan(ws5, r, 1, 4, "ИТОГО НА 100 ВИДЕО/ДЕНЬ", bold=True, fc=WHITE, bg=INK, align="right", border=False, indent=1)
cl(ws5, r, 5, "≈19 мин/видео", bold=True, bg=INK, fc=WHITE, align="center")
cl(ws5, r, 6, "≈30 ч / 14 чел", bold=True, bg=INK, fc=WHITE, align="center")
mspan(ws5, r, 7, 8, "При полной автоматизации войсовера и субтитров", italic=True, fc=WHITE, bg=INK, sz=8.5, border=False, indent=1)
rh(ws5, r, 24)

freeze_below_header(ws5, 5)
print("Sheet 5 (Pipeline) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 6 — EXECUTION: TEAM (Layer 4)
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("👥 06 Team")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = L_EXEC
wid(ws6, [24, 9, 42, 30, 14, 16])

r = header_block(
    ws6, 1, 6,
    "LAYER 4 · EXECUTION",
    "👥  КОМАНДА И РОЛИ",
    "Минимальный состав, необходимый для устойчивых 100 видео/день.",
    L_EXEC,
)

r = section(ws6, r, 1, 6, "  СОСТАВ КОМАНДЫ", L_EXEC)
heads = ["Роль", "Кол-во", "Зона ответственности", "Инструменты", "Ставка/мес ($)", "Итого/мес ($)"]
for ci, h in enumerate(heads, 1):
    hd(ws6, r, ci, h)
rh(ws6, r, 24)
r += 1
team_start = r
team = [
    ("Контент-директор", 1, "Стратегия, editorial plan, утверждение форматов, QC-шлюз", "Notion, Metricool", 3000),
    ("Prompt-инженер", 2, "Промпты для скриптов/голоса/видео, шаблоны автоматизации", "Claude, GPT-4o, ElevenLabs", 1500),
    ("Видеоредактор", 4, "Монтаж, субтитры, обложки — 25 видео/день на человека", "CapCut Pro, Canva", 800),
    ("Контент-стратег", 2, "Контент-план, рубрики, A/B хуков, промо-слот", "Notion, Miro, Sheets", 1200),
    ("Тренд-ресёрчер", 2, "Мониторинг тем, конкуренты, ключевые слова", "GPT-4o, Google Trends", 700),
    ("Менеджер аккаунтов", 2, "Публикация, расписание, ответы на комментарии", "Later, Buffer", 800),
    ("Аналитик", 1, "Отчётность, kill/scale-решения, оптимизация воронки", "Metricool, GA4", 1500),
]
for i, (role, cnt, duties, tools, sal) in enumerate(team):
    rr = team_start + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws6, rr, 1, role, bold=True, bg=bg, fc=TEXT)
    cl(ws6, rr, 2, cnt, bg=bg, align="center")
    cl(ws6, rr, 3, duties, bg=bg, fc=TEXT, sz=8.8)
    cl(ws6, rr, 4, tools, italic=True, bg=bg, fc=MUTED, sz=8.5)
    cl(ws6, rr, 5, sal, bg=bg, align="center", nf='"$"#,##0')
    cl(ws6, rr, 6, f"=B{rr}*E{rr}", bg=ACCENT_BG, align="center", bold=True, fc=ACCENT, nf='"$"#,##0')
    rh(ws6, rr, 38)
team_end = team_start + len(team) - 1
r = team_end + 1

mspan(ws6, r, 1, 4, "ИТОГО КОМАНДА / МЕСЯЦ", bold=True, fc=WHITE, bg=INK, align="right", border=False, indent=1)
cl(ws6, r, 5, f"=SUM(B{team_start}:B{team_end})", bold=True, bg=INK, fc=WHITE, align="center")
cl(ws6, r, 6, f"=SUM(F{team_start}:F{team_end})", bold=True, bg=INK, fc=WHITE, align="center", nf='"$"#,##0')
TEAM_TOTAL_CELL = f"'👥 06 Team'!F{r}"
rh(ws6, r, 24)
r += 2

notes6 = [
    "Ставки ориентированы на удалённых специалистов СНГ/Азии, не США/Европа.",
    "Prompt-инженер и тренд-ресёрчер могут совмещаться на фазе 0–1 (до 60 видео/день).",
    "Фаза 0–1 (тест): команда из 7 человек; полный найм — к фазе 2 (см. 01 Strategy).",
    "Видеомонтаж частично выносится на аутсорс (Fiverr/Toptal) для снижения фикс-затрат.",
]
r = section(ws6, r, 1, 6, "  ПРИМЕЧАНИЯ", SLATE)
for i, t in enumerate(notes6):
    r = note(ws6, r, 1, 6, "•  " + t, h=20, bg=ROW_ALT if i % 2 == 0 else WHITE)

freeze_below_header(ws6, 5)
print("Sheet 6 (Team) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 7 — EXECUTION: AI STACK (Layer 4)
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("🛠️ 07 AI Stack")
ws7.sheet_view.showGridLines = False
ws7.sheet_properties.tabColor = L_EXEC
wid(ws7, [16, 26, 34, 14, 24])

r = header_block(
    ws7, 1, 5,
    "LAYER 4 · EXECUTION",
    "🛠️  AI-СТЕК И АВТОМАТИЗАЦИЯ",
    "Технологический слой, который позволяет 14 людям производить 100 видео/день.",
    L_EXEC,
)

r = section(ws7, r, 1, 5, "  ИНСТРУМЕНТЫ ПО КАТЕГОРИЯМ", L_EXEC)
heads = ["Категория", "Инструмент", "Назначение", "Цена/мес ($)", "Примечание"]
for ci, h in enumerate(heads, 1):
    hd(ws7, r, ci, h)
rh(ws7, r, 24)
r += 1
tool_start = r
tools = [
    ("Скрипты & идеи", "ChatGPT Plus × 5", "Генерация скриптов, тренд-ресёрч, хуки", 100, "Batch-промпты"),
    ("Скрипты & идеи", "Claude Pro × 3", "Длинные скрипты, анализ конкурентов", 60, "Лучше для связных текстов"),
    ("Голос", "ElevenLabs Creator", "AI-войсовер, до 100 треков/день", 99, "Резерв: Business при нехватке лимита"),
    ("Видео", "CapCut Pro × 10", "Монтаж, авто-субтитры, шаблоны", 120, "Ключевой инструмент пайплайна"),
    ("Видео", "HeyGen Business", "AI-аватар, лицо ведущего", 480, "Альтернатива: Synthesia"),
    ("Видео", "Runway Gen-3 Standard", "AI B-roll, визуальные эффекты", 95, "Альтернатива: Pika Labs"),
    ("Дизайн", "Canva Pro × 5", "Обложки, оверлеи, брендинг", 65, "Шаблоны на все 15 аккаунтов"),
    ("Дизайн", "Midjourney Pro × 3", "AI-изображения для превью", 90, "Fast hours для batch-режима"),
    ("Публикация", "Later Growth", "Расписание постов, 5 аккаунтов", 80, "—"),
    ("Публикация", "Buffer Essentials", "Доп. 10 аккаунтов, аналитика", 100, "10 каналов"),
    ("Аналитика", "Metricool Business", "Единый дашборд на 15 аккаунтов", 119, "Ключевой инструмент 08 KPI Dashboard"),
    ("Аналитика", "Social Blade Pro", "Мониторинг роста аккаунтов", 20, "Конкурентная разведка"),
    ("Инфраструктура", "VPN/прокси × 15", "Безопасность мультиаккаунтинга", 100, "BrightData / Smartproxy"),
    ("Инфраструктура", "Google Workspace", "Хранение асетов, 2TB", 36, "Командная работа"),
    ("Инфраструктура", "Notion Team", "Контент-календарь, база знаний", 20, "CRM для контента"),
]
prev_cat = None
for i, (cat, tool, purp, price, note_) in enumerate(tools):
    rr = tool_start + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cat_val = cat if cat != prev_cat else ""
    cl(ws7, rr, 1, cat_val, bold=True, bg=bg, fc=L_EXEC, sz=9)
    cl(ws7, rr, 2, tool, bold=True, bg=bg, fc=TEXT, sz=9)
    cl(ws7, rr, 3, purp, bg=bg, fc=TEXT, sz=8.8)
    cl(ws7, rr, 4, price, bg=ACCENT_BG, align="center", fc=ACCENT, bold=True, nf='"$"#,##0')
    cl(ws7, rr, 5, note_, italic=True, bg=bg, fc=MUTED, sz=8.3)
    rh(ws7, rr, 26)
    prev_cat = cat
tool_end = tool_start + len(tools) - 1
r = tool_end + 1

mspan(ws7, r, 1, 3, "ИТОГО AI-ИНСТРУМЕНТЫ / МЕСЯЦ", bold=True, fc=WHITE, bg=INK, align="right", border=False, indent=1)
cl(ws7, r, 4, f"=SUM(D{tool_start}:D{tool_end})", bold=True, bg=INK, fc=WHITE, align="center", nf='"$"#,##0')
mspan(ws7, r, 5, 5, "Без платного трафика", italic=True, fc=WHITE, bg=INK, sz=8.5, border=False, indent=1)
TOOLS_TOTAL_CELL = f"'🛠️ 07 AI Stack'!D{r}"
TOOLS_TOTAL_ROW = r
rh(ws7, r, 24)
r += 2

r = section(ws7, r, 1, 5, "  ОБЩИЙ БЮДЖЕТ ИСПОЛНЕНИЯ (cross-sheet)", SLATE)
mspan(ws7, r, 1, 3, "Команда (см. 06 Team)", bold=True, bg=ROW_ALT, fc=TEXT, align="left")
cl(ws7, r, 4, f"={TEAM_TOTAL_CELL}", bold=True, bg=ROW_ALT, fc=ACCENT, align="center", nf='"$"#,##0')
cl(ws7, r, 5, "", bg=ROW_ALT)
rh(ws7, r, 22)
r += 1
mspan(ws7, r, 1, 3, "AI-инструменты (выше)", bold=True, bg=WHITE, fc=TEXT, align="left")
cl(ws7, r, 4, f"={TOOLS_TOTAL_CELL}", bold=True, bg=WHITE, fc=ACCENT, align="center", nf='"$"#,##0')
cl(ws7, r, 5, "", bg=WHITE)
rh(ws7, r, 22)
r += 1
mspan(ws7, r, 1, 3, "ИТОГО ЗАПУСК ИСПОЛНЕНИЯ / МЕСЯЦ", bold=True, fc=WHITE, bg=INK, align="right", border=False, indent=1, sz=10.5)
cl(ws7, r, 4, f"={TEAM_TOTAL_CELL}+{TOOLS_TOTAL_CELL}", bold=True, bg=INK, fc=WHITE, align="center", nf='"$"#,##0', sz=10.5)
cl(ws7, r, 5, "", bg=INK)
rh(ws7, r, 26)

freeze_below_header(ws7, 5)
print("Sheet 7 (AI Stack) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 8 — KPI & ANALYTICS (Layer 5)
# ════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("📈 08 KPI Dashboard")
ws8.sheet_view.showGridLines = False
ws8.sheet_properties.tabColor = L_KPI
wid(ws8, [30, 14, 14, 14, 26, 13])

r = header_block(
    ws8, 1, 6,
    "LAYER 5 · KPI & ANALYTICS",
    "📈  ДАШБОРД — ПРОСМОТРЫ · ENGAGEMENT · РОСТ · УДЕРЖАНИЕ · КОНВЕРСИИ",
    "Если метрика не в этом списке — она не двигает решения kill/scale (см. 01 Strategy, принцип 3).",
    L_KPI,
)

r = section(ws8, r, 1, 6, "  МЕТРИКИ И ЦЕЛИ ПО МЕСЯЦАМ", L_KPI)
hd(ws8, r, 1, "Метрика"); hd(ws8, r, 2, "Мес. 1"); hd(ws8, r, 3, "Мес. 3"); hd(ws8, r, 4, "Мес. 6 (цель)")
hd(ws8, r, 5, "Как считается / инструмент"); hd(ws8, r, 6, "Частота")
rh(ws8, r, 24)
r += 1

def kpi_section(ws, row, label):
    section(ws, row, 1, 6, "  " + label, SLATE, h=20)
    return row + 1

kpi_groups = [
    ("1 · ПРОСМОТРЫ", [
        ("Просмотры/месяц, все аккаунты", "5 000 000", "40 000 000", "100 000 000", "Сумма Reels-просмотров · Metricool / Insights", "Еженедельно"),
        ("Среднее просмотров на видео", "1 667", "13 333", "33 333", "Views ÷ опубликованные видео за период", "Еженедельно"),
        ("Доля видео с виральным охватом (>10× подписчиков)", "1–2%", "5–8%", "10%+", "Ручной мониторинг по 09 Content Bank", "Еженедельно"),
    ]),
    ("2 · ENGAGEMENT RATE", [
        ("Engagement Rate, средний", "3,0%", "4,0%", "5,0%", "(Лайки+Комм.+Сейвы+Шеры) ÷ Просмотры", "Еженедельно"),
        ("Sends per reach (репосты в директ)", "0,3%", "0,6%", "1,0%+", "Метрика Instagram Insights «Sends»", "Еженедельно"),
        ("Saves per reach (сохранения)", "0,5%", "1,0%", "1,5%+", "Метрика Instagram Insights «Saves»", "Еженедельно"),
        ("Глубина комментариев (% тредов >1 ответа)", "10%", "20%", "30%", "Ручная выборка топ-20 видео/нед.", "Еженедельно"),
    ]),
    ("3 · РОСТ АККАУНТОВ", [
        ("Суммарные подписчики, все аккаунты", "10 000", "600 000", "4 650 000", "Social Blade / Insights, сумма по 02 Accounts", "Еженедельно"),
        ("Средний прирост подписчиков/аккаунт/мес", "—", "30 000", "50 000", "Δ подписчиков за период ÷ кол-во аккаунтов", "Еженедельно"),
        ("Non-follower reach % (охват от не-подписчиков)", "40%", "55%", "65%+", "Instagram Insights — канареечная метрика алгоритма", "Ежедневно"),
    ]),
    ("4 · УДЕРЖАНИЕ (RETENTION)", [
        ("Среднее время просмотра, % от длины", "35%", "50%", "60%+", "Instagram Insights «Average watch time» ÷ длина ролика", "Еженедельно"),
        ("Completion rate (досмотры до конца)", "20%", "35%", "45%+", "Instagram Insights «Plays at 100%»", "Еженедельно"),
        ("Replay rate (повторные просмотры)", "5%", "10%", "15%+", "Instagram Insights «Replays» ÷ Reach", "Еженедельно"),
        ("3-секундный drop-off", "55%", "40%", "<30%", "Доля ушедших до 3-й секунды — диагностика хука", "Еженедельно"),
    ]),
    ("5 · КОНВЕРСИИ", [
        ("Переходы в профиль → подписка, %", "8%", "12%", "15%", "Profile visits → Follows, Instagram Insights", "Еженедельно"),
        ("Переходы на сайт/лендинг/мес", "500", "10 000", "50 000", "UTM-метки + Google Analytics", "Еженедельно"),
        ("Лиды (заявки)/месяц", "50", "700", "3 000", "CRM продукта", "Ежемесячно"),
        ("CPL — стоимость лида", "$400", "$28", "$7", "Бюджет исполнения ÷ лиды (см. 07 AI Stack)", "Ежемесячно"),
    ]),
]
for label, rows in kpi_groups:
    r = kpi_section(ws8, r, label)
    for i, (metric, m1, m3, m6, tool, freq) in enumerate(rows):
        rr = r + i
        bg = ROW_ALT if i % 2 == 0 else WHITE
        cl(ws8, rr, 1, metric, bold=True, bg=bg, fc=TEXT, sz=9)
        cl(ws8, rr, 2, m1, bg=ACCENT_BG, fc=ACCENT, align="center", sz=9)
        cl(ws8, rr, 3, m3, bg=WARN_BG, fc=WARN, align="center", sz=9, bold=True)
        cl(ws8, rr, 4, m6, bg=GOOD_BG, fc=GOOD, align="center", sz=9, bold=True)
        cl(ws8, rr, 5, tool, italic=True, bg=bg, fc=MUTED, sz=8.3)
        cl(ws8, rr, 6, freq, bg=bg, align="center", sz=8.3)
        rh(ws8, rr, 22)
    r += len(rows)

freeze_below_header(ws8, 5)
print("Sheet 8 (KPI Dashboard) built")

# ════════════════════════════════════════════════════════════════════
# SHEET 9 — SHORT-FORM CONTENT BANK — 100 SCRIPTS (Layer 3 deliverable)
# ════════════════════════════════════════════════════════════════════
PRINC_META = {
    "ME": ("Mere Exposure",        "D9E6FB", "1B4DA6"),
    "CT": ("Curiosity Trap",       "FBE8C7", "8A5A00"),
    "IS": ("Identity Selling",     "ECE0F8", "5B3489"),
    "AC": ("Authority",            "D8F0EF", "0B6360"),
    "SS": ("Social Sharing",       "FBE2E0", "9C2A24"),
    "ST": ("Storytelling",         "DFF3E8", "10663F"),
}

ME = [
    ("AI-новости", "Рубрика «AI за 60 секунд», выпуск #24 — сегодня разбираем Gemini",
     "Та же заставка/музыка/ведущий, что и в прошлых 23 выпусках → 1 новость → итог одной фразой",
     "Не вирусится сама по себе, но удерживает постоянную аудиторию, которая досматривает на автомате узнавания",
     "Высокий repeat-view, комментарии-предсказания «сегодня точно про OpenAI»"),
    ("Туториал", "Понедельничный AI-лайфхак: сегодня — Notion AI",
     "Фирменный текстовый баннер «понедельничный лайфхак» открывает каждый ролик серии → 1 приём за 20 сек → призыв сохранить",
     "Снижает % отписок за счёт предсказуемого ритма публикаций",
     "Стабильные сейвы по понедельникам, рост узнаваемости профиля в ленте"),
    ("Сравнение", "Битва инструментов, раунд 9: Claude vs ChatGPT",
     "Узнаваемый «счётный» оверлей раунда → задача → результат → счёт обновляется в кадре",
     "Серийный формат вызывает ожидание следующего раунда, комментарии-ставки",
     "Комментарии-прогнозы, возврат зрителей к следующему эпизоду"),
    ("Шоукейс", "Пятничный AI-арт: Midjourney рисует то, что попросили в комментариях",
     "Тот же интро-джингл «пятничный AI-арт» → промпт из комментариев → результат крупным планом",
     "Зрители возвращаются ради своей идеи в следующем выпуске",
     "Высокая вовлечённость в комментарии-заявки, стабильный охват по подписчикам"),
    ("Продуктивность", "Чек-лист понедельника: 3 AI-инструмента на неделю",
     "Фирменная палитра и шрифт чек-листа повторяются каждую неделю → 3 пункта → итоговая карточка",
     "Узнаваемая карточка располагает к сохранению как «еженедельный дайджест»",
     "Сохранения выше среднего, комментарии с собственными списками"),
    ("AI-новости", "Голос за кадром тот же, что и всегда — у Gemini обновили память",
     "Постоянный голос и темп озвучки на каждом видео → факт → вывод",
     "Узнаваемый голос ассоциируется с надёжным источником, а не разовой сенсацией",
     "Снижение оттока подписчиков, стабильный CTR в Stories"),
    ("Туториал", "Серия «1 минута — 1 функция»: сегодня Gamma",
     "Единый формат таймера 60 секунд в углу экрана на каждом видео серии → демонстрация → таймер на нуле",
     "Таймер создаёт лёгкий саспенс, формат держит привычкой, не открытием",
     "Высокая частота повторных просмотров — зрители пересматривают, чтобы успеть повторить"),
    ("Сравнение", "Таблица недели: 5 AI-инструментов для почты",
     "Одна и та же сетка-таблица как визуальный шаблон рубрики → заполнение по ходу видео → финальная сводная таблица",
     "Таблица шерится как референс, узнаваемость формата облегчает решение «смотреть/нет»",
     "Высокие сейвы, репосты как «шпаргалка»"),
    ("Юмор", "Рубрика «AI облажался» #11",
     "Узнаваемый звуковой сигнал ошибки в начале каждого выпуска → провал AI → реакция в кадре",
     "Узнаваемый звук создаёт условный рефлекс «сейчас будет смешно»",
     "Высокий watch-through благодаря привычному ритму юмора"),
    ("Промо", "История клиента №7: как [продукт] сэкономил время",
     "Тот же шаблон «До / После» с таймером экономии часов в каждом промо-видео",
     "Узнаваемый формат повышает доверие к новому кейсу за счёт прошлых выпусков",
     "Стабильная конверсия в переходы на сайт"),
    ("AI для бизнеса", "Кейс недели: ROI от AI в реальной компании",
     "Фирменная инфографика ROI в одном стиле каждую неделю → цифры → итог",
     "Сериальность повышает доверие B2B-аудитории к формату отчётности",
     "Сохранения как референс для презентаций"),
    ("Туториал", "Урок #32 из бесплатного AI-курса в этом профиле",
     "Нумерация уроков с одинаковым шаблоном слайда «Урок №» → объяснение → задание в подписи",
     "Эффект учебного сериала удерживает на канале, а не делится вовне",
     "Высокая частота возвратов, рост сохранённых сборников"),
    ("AI-новости", "Сводка за неделю: 3 релиза, которые вы пропустили",
     "Один и тот же формат «3 новости подряд» с разделителями-таймкодами",
     "Предсказуемая структура снижает порог досмотра до конца",
     "Высокий completion rate за счёт привычного темпа"),
    ("Продуктивность", "Утренний ритуал с AI: серия продолжается",
     "Один и тот же визуальный сеттинг рабочего стола в каждом видео серии",
     "Узнаваемая «утренняя атмосфера» формирует привычку смотреть в одно время",
     "Рост просмотров в определённый час дня, паттерн постоянной аудитории"),
    ("Сравнение", "Голосование недели: какой AI-инструмент тестируем дальше",
     "Фирменный опрос в Stories перед роликом → ролик по итогам со знакомой подачей",
     "Чувство соучастия аудитории в формате повышает лояльность и регулярный просмотр",
     "Высокая активность в Stories, переход в комментарии под видео"),
    ("Шоукейс", "AI-аватар снова здесь — досматривай до сюрприза",
     "Один и тот же HeyGen-аватар ведёт уже n-й выпуск рубрики → демонстрация → фирменный финальный жест",
     "Персонаж-ведущий формирует парасоциальную привязанность, типичную для серийного контента",
     "Стабильный рост подписки именно на личность аватара, а не на разовый ролик"),
]

CT = [
    ("Туториал", "Есть настройка ChatGPT, которую не включают 99% — досмотри до конца",
     "Hook не называет настройку → Problem: долгие неточные ответы → Story: ищем скрытый пункт меню → Resolution: настройка раскрывается на 27-й секунде",
     "Прямой open loop держит до конца, что повышает вес ролика по watch time",
     "Очень высокий completion rate, мало ранних уходов"),
    ("Сравнение", "Один из этих AI-инструментов тайно собирает больше данных, чем вы думаете",
     "Не называем какой → сравниваем интерфейсы → раскрываем разницу в политике данных в payoff",
     "Тревожный open loop про приватность вызывает желание узнать, какой именно",
     "Высокие комментарии «который из них?», споры в обсуждении"),
    ("Новость", "OpenAI выпустила фичу, но включили её не всем — проверь, есть ли она у тебя",
     "Не раскрываем фичу сразу → контекст релиза → как проверить наличие → демонстрация",
     "Личная неопределённость «есть ли это у меня» удерживает каждого зрителя индивидуально",
     "Высокие сейвы «чтобы потом проверить», переходы в настройки сразу после просмотра"),
    ("Шоукейс", "Это не Photoshop — и не Midjourney тоже",
     "Показываем результат без подписи инструмента → демонстрация процесса → раскрытие имени на 25-й секунде",
     "Загадка «что это» провоцирует комментарии-догадки до раскрытия ответа",
     "Высокая комментарийная активность с вариантами ответа"),
    ("Продуктивность", "Один промпт убирает половину твоей работы в Excel — но он не для всех версий",
     "Условие «не для всех» создаёт разрыв → демонстрация промпта → уточнение совместимости в конце",
     "Условность интригует и заставляет смотреть, чтобы понять «подхожу ли я»",
     "Высокая досматриваемость, комментарии с вопросами о версии"),
    ("Юмор", "ИИ почти угадал, но в последний момент всё испортил",
     "Показываем процесс генерации без финального кадра → нарастающее ожидание → провал в payoff",
     "Комедийный саспенс вокруг «угадает или нет» держит внимание до конца",
     "Высокий completion, частые повторные просмотры момента провала"),
    ("Туториал", "Этот промпт работает только если написать его именно так",
     "Подчёркиваем точность формулировки как условие → ошибочный и верный вариант → точный текст в payoff",
     "Требование точности создаёт необходимость пересмотра и сохранения",
     "Высокие сейвы, повторные просмотры для копирования точного текста"),
    ("Сравнение", "Спросили один и тот же вопрос у трёх AI — один ответ был неожиданно честным",
     "Не говорим какой → задаём вопрос трём моделям → раскрываем «честный» ответ в конце",
     "Любопытство «кто оказался честным» удерживает до payoff",
     "Высокие комментарии с обсуждением, кто угадал правильно"),
    ("Новость", "В обновлении есть строчка, которую почти никто не заметил",
     "Контекст релиза → перечисление заметных фич → раскрытие незаметной строчки как главного твиста",
     "Эффект «скрытой детали» провоцирует пересмотр официального анонса другими зрителями",
     "Высокий completion rate, комментарии «а я и не заметил»"),
    ("Шоукейс", "Я попросил AI закончить мою мысль — ответ удивил даже меня",
     "Демонстрируем начало запроса → пауза перед ответом → реакция в кадре как часть payoff",
     "Личная реакция автора усиливает эффект ожидания для зрителя",
     "Высокая досматриваемость ради финальной реакции"),
    ("Продуктивность", "Если твой календарь выглядит так — тебе не хватает одной настройки",
     "Показываем хаотичный календарь → намекаем на решение → раскрываем настройку в конце",
     "Самоидентификация по картинке календаря держит внимание тех, кто узнал себя",
     "Высокие сейвы как «инструкция на потом»"),
    ("AI для бизнеса", "Цифра в этом кейсе вас удивит — ROI не там, где ожидаешь",
     "Задаём контекст кейса → нарастание цифр → неожиданный итоговый показатель",
     "Контринтуитивный результат заставляет досмотреть, чтобы узнать конкретное число",
     "Высокий completion, сохранения для презентаций"),
    ("Туториал", "В Notion AI есть функция, спрятанная в трёх кликах от всех остальных",
     "Подчёркиваем «спрятанность» → путь из трёх кликов в реальном времени → демонстрация функции",
     "Зрители досматривают, чтобы получить точный путь, не запоминая на слух",
     "Высокая досматриваемость, репост теми, кто уже использует инструмент"),
    ("Сравнение", "Один из этих инструментов скоро станет платным — угадай, какой",
     "Сравниваем функциональность вслепую → намёк на смену модели монетизации → раскрытие в payoff",
     "Финансовая интрига «успею ли я» мотивирует досмотр и сохранение",
     "Высокие сейвы «чтобы успеть», комментарии с предположениями"),
    ("Новость", "Это не баг — это специально скрытая функция нового релиза",
     "Показываем «странное» поведение интерфейса → переворот ожидания → объяснение как фичи",
     "Переворот ожидания «баг или фича» — классический crowd-triggering твист",
     "Высокий completion rate, споры в комментариях до объяснения"),
    ("Юмор", "Я не скажу, что попросил — угадай по результату",
     "Показываем только результат генерации → зрители гадают в комментариях → промпт раскрывается в конце",
     "Игровой формат угадывания стимулирует комментирование до раскрытия ответа",
     "Высокая комментарийная активность, рост охвата за счёт обсуждения"),
    ("Продуктивность", "Эта мелочь в Google Calendar экономит час в неделю, но её выключают по умолчанию",
     "Указываем на дефолтное состояние «выключено» как разрыв → шаги включения → итоговый эффект",
     "Конкретная цифра экономии плюс ощущение упущенной выгоды удерживает внимание",
     "Высокие сейвы, переходы в настройки сразу после просмотра"),
]

IS = [
    ("Продуктивность", "Если ты ещё закрываешь отчёты вручную — это видео для тебя",
     "Обозначаем «отстающую» идентичность → демонстрация AI-решения → новый статус «продвинутого» в payoff",
     "Зритель делится, чтобы дистанцироваться от «отстающей» идентичности и показать новый статус",
     "Высокие репосты коллегам с подписью «нам это нужно»"),
    ("Сравнение", "AI-пользователи топ-1% никогда не пишут промпт вот так",
     "Контраст «обычный» vs «продвинутый» промпт → разница в результате → приглашение войти в группу «1%»",
     "Аспирационная идентичность мотивирует сохранять и подражать «инсайдерам»",
     "Высокие сейвы, попытки повторить «продвинутый» промпт в комментариях"),
    ("Туториал", "Это для тех, кто называет себя организованным, но живёт в хаосе вкладок",
     "Узнаваемая самоирония → демонстрация AI-инструмента наведения порядка → финальный «чистый» экран",
     "Самоидентификация с лёгкой самоиронией провоцирует тег друзей «это про нас»",
     "Высокие отметки друзей в комментариях"),
    ("AI для бизнеса", "Менеджеры, которые ещё не используют AI в отчётах, скоро будут не нужны",
     "Тревожная идентичность «устаревающего специалиста» → конкретное решение → идентичность «незаменимого»",
     "Страх профессионального устаревания — сильный мотиватор пересылки коллегам",
     "Высокие репосты внутри рабочих чатов, сохранения «к разговору с командой»"),
    ("Юмор", "POV: ты всё ещё гуглишь, пока коллега уже спросил у AI",
     "Комедийный контраст двух идентичностей в одном офисе → демонстрация скорости AI-метода → ирония в payoff",
     "Узнаваемая офисная ситуация провоцирует тег конкретного коллеги",
     "Высокая активность тегов в комментариях, шеры в рабочие чаты"),
    ("Продуктивность", "Фрилансеры, которые ценят своё время, делают именно так",
     "Аспирационная идентичность «ценящего время» → конкретный AI-приём → результат в часах",
     "Зритель шерит, чтобы подтвердить принадлежность к группе «ценящих время»",
     "Высокие сохранения, репосты во фриланс-сообщества"),
    ("Туториал", "Если ты студент и ещё не настроил это — ты усложняешь себе жизнь зря",
     "Идентичность «студента, который усложняет себе жизнь» → пошаговая настройка → облегчённый процесс в финале",
     "Студенческая аудитория активно делится лайфхаками внутри учебных групп",
     "Высокие репосты в групповые чаты, сохранения перед сессией"),
    ("Сравнение", "Опытные маркетологи выбирают между этими двумя иначе, чем новички",
     "Контраст «новичок» vs «опытный» в выборе инструмента → критерий выбора → приглашение в группу «опытных»",
     "Профессиональная идентичность стимулирует пересылку внутри маркетинговых команд",
     "Высокие сохранения как референс для команды"),
    ("Юмор", "Это видео не для тех, кто ещё печатает письма с нуля — для остальных",
     "Лёгкое исключение «непродвинутых» из аудитории → демонстрация AI-шаблона → ирония в payoff",
     "Лёгкий троллинг идентичности провоцирует пересылку и как шутку, и как подсказку",
     "Высокие репосты с подписью «это тебе»"),
    ("AI-новости", "Те, кто следит за AI каждый день, уже знают эту новость",
     "Идентичность «осведомлённого инсайдера» → подача новости «для своих» → деталь, неизвестная массовой аудитории",
     "Зритель делится, чтобы подтвердить статус «следящего за трендами»",
     "Высокие репосты как демонстрация осведомлённости"),
    ("Продуктивность", "Руководители, у которых нет времени на хаос, делают это в начале недели",
     "Идентичность «организованного руководителя» → конкретный понедельничный AI-ритуал → итог: чистая неделя",
     "Управленческая аудитория делится внутри руководящих чатов как стандарт",
     "Высокие сохранения, репосты в управленческие сообщества"),
    ("Туториал", "Разработчики, которые ценят свой код, не пишут boilerplate руками в 2026",
     "Профессиональная идентичность «ценящего качество кода» → демонстрация AI-ассистента → ускоренный результат",
     "Технически подкованная аудитория делится в dev-комьюнити как маркер актуальности",
     "Высокие репосты в dev-чаты и каналы"),
    ("AI для бизнеса", "Компании, которые масштабируются без найма, делают именно это",
     "Аспирационная бизнес-идентичность «эффективных» → кейс автоматизации → результат в цифрах",
     "Предприниматели делятся, чтобы ассоциировать себя с «эффективными» компаниями",
     "Высокие сохранения, пересылка инвесторам и партнёрам"),
    ("Сравнение", "Дизайнеры, которые не боятся AI, выбирают этот инструмент по другой причине",
     "Идентичность «небоящегося дизайнера» против скептиков → объяснение критерия → приглашение в группу",
     "Профессиональная гордость стимулирует репост в дизайн-сообщества",
     "Высокие сохранения, обсуждение в комментариях с про- и контра-аргументами"),
    ("Продуктивность", "Это для тех, кто устал быть «всегда онлайн» по работе",
     "Идентичность уставшего от переработок → AI-решение для границ → новая идентичность «человека с балансом»",
     "Эмоциональный отклик про баланс жизни и работы провоцирует репост близким",
     "Высокие репосты с личными комментариями о выгорании"),
    ("Туториал", "Копирайтеры нового поколения больше не боятся чистого листа — вот почему",
     "Идентичность «нового поколения» против страха чистого листа → AI-приём для старта текста → итог",
     "Профессиональная идентичность мотивирует шерить внутри копирайтерских комьюнити",
     "Высокие сохранения, обсуждения техник в комментариях"),
    ("Промо", "Студенты, которые внедрили AI первыми, уже обогнали поток",
     "Идентичность «опередившего поток» → конкретный кейс выпускника → приглашение присоединиться",
     "Социальное доказательство плюс страх отстать мотивирует репост в учебные чаты",
     "Высокие переходы по ссылке, репосты потенциальным студентам"),
]

AC = [
    ("Сравнение", "Я протестировал 47 AI-инструментов за месяц — вот топ-3 для работы с текстом",
     "Скринкаст интерфейса вместо лица как доказательство объёма тестирования в первом кадре → 3 инструмента с цифрами → итоговый рейтинг",
     "Доверие к объёму экспертизы повышает шанс сохранения как авторитетного источника",
     "Высокие сохранения, низкий ранний отток благодаря мгновенному доверию"),
    ("Туториал", "8 лет в маркетинге — и это первый AI-инструмент, который реально меняет процесс",
     "Конкретный профессиональный стаж как credential → демонстрация инструмента → личная оценка эксперта",
     "Профессиональный авторитет повышает вес рекомендации для целевой аудитории",
     "Высокий completion rate, доверительные комментарии с вопросами"),
    ("Новость", "Я был на презентации релиза — вот что не попало в официальный анонс",
     "Личное присутствие как доказательство первоисточника → детали, не описанные в пресс-релизе → вывод",
     "Эксклюзивность инсайдерской информации повышает доверие и желание поделиться «из первых рук»",
     "Высокие репосты с пометкой «из первых рук»"),
    ("AI для бизнеса", "Цифры из реального проекта на 50 000 строк данных — не из демо",
     "Указание реального масштаба данных в первом кадре → процесс на реальном датасете → измеримый итог",
     "Достоверность реального кейса повышает доверие B2B-аудитории сильнее синтетического демо",
     "Высокие сохранения, переходы в профиль за консультацией"),
    ("Шоукейс", "Скринкаст, не постановка: я не останавливал запись ни разу",
     "Явное заявление о непрерывной записи как доказательство честности → процесс целиком → результат",
     "Подчёркнутая прозрачность повышает доверие в нише, где много постановочных демо",
     "Высокий completion rate, комментарии с благодарностью за честность"),
    ("Туториал", "Сертифицированный тренер по Notion показывает то, что не входит в официальный курс",
     "Профессиональная сертификация как credential → демонстрация продвинутого приёма → личная рекомендация",
     "Экспертный статус повышает доверие, особенно у нишевой профессиональной аудитории",
     "Высокие сохранения как обучающий материал"),
    ("Сравнение", "После 200 часов работы в обоих инструментах — вот честный вердикт",
     "Конкретное количество часов как доказательство глубины опыта → сравнение по реальным сценариям → итог",
     "Объём личного опыта повышает вес мнения сильнее, чем у разовых обзоров",
     "Высокие сохранения, доверительные комментарии-вопросы по нюансам"),
    ("AI-новости", "Я читаю changelog каждого релиза — вот что реально важно из вчерашнего обновления",
     "Регулярность мониторинга как credential → фильтрация важного из списка изменений → вывод для пользователя",
     "Экспертная фильтрация информации экономит время зрителю, повышая шанс сохранения и подписки",
     "Высокий completion rate, рост подписки как на «надёжный фильтр новостей»"),
    ("Продуктивность", "Я веду команду из 12 человек на AI-автоматизации — вот стек целиком",
     "Управленческий опыт и масштаб команды как credential → обзор реального стека → результат по времени команды",
     "Реальный управленческий кейс убедительнее теоретических советов для руководителей",
     "Высокие сохранения, переходы в профиль за консультацией"),
    ("Туториал", "Бывший инженер крупной AI-лаборатории объясняет, как на самом деле работает промпт-инжиниринг",
     "Профессиональный бэкграунд как credential → разбор механики на конкретном примере → практический вывод",
     "Инсайдерский профессиональный опыт резко повышает воспринимаемую достоверность",
     "Высокие сохранения, репосты как «экспертное мнение»"),
    ("Сравнение", "Я тестировал AI-инструменты в трёх агентствах — везде разные победители",
     "Множественный контекст применения как credential → сравнение по типу агентства → нюансированный вывод",
     "Нюансированная экспертиза (не «всем подходит одно») повышает доверие искушённой аудитории",
     "Высокие сохранения, обсуждение в комментариях своего контекста"),
    ("AI для бизнеса", "Финансовый директор объясняет, как считать ROI от AI правильно",
     "Профессиональная роль как credential → разбор частой ошибки в расчёте → правильная формула",
     "Финансовая экспертиза резко повышает доверие к цифрам ROI у бизнес-аудитории",
     "Высокие сохранения как референс для собственных расчётов"),
    ("Туториал", "10 000 промптов протестировано — вот формула, которая работает чаще всего",
     "Объём тестирования как credential → разбор формулы по частям → пример применения",
     "Статистическая основательность сильнее единичного личного опыта",
     "Высокие сохранения, попытки применить формулу в комментариях"),
    ("Шоукейс", "Профессиональный видеомонтажёр: вот где AI ещё не заменит человека",
     "Профессиональный статус как credential для честного контринтуитивного вывода → примеры ограничений → баланс мнения",
     "Честность эксперта о пределах технологии повышает доверие сильнее, чем хайп",
     "Высокие сохранения, доверительные обсуждения в комментариях"),
    ("Продуктивность", "Коуч по продуктивности 6 лет — вот единственный AI-инструмент, который я реально использую",
     "Профессиональный опыт как credential → честный отбор «единственного» инструмента из множества → причина выбора",
     "Избирательность эксперта повышает доверие через простоту выбора",
     "Высокие сохранения, рост подписки на профиль эксперта"),
    ("AI-новости", "Я внедрял этот API в день релиза — вот как он работает на практике",
     "Техническая практика день-в-день как credential → демонстрация реальной интеграции → практический вывод",
     "Техническая глубина в день релиза резко повышает доверие разработчиков",
     "Высокие репосты в dev-сообщества, сохранения как техническая заметка"),
    ("Сравнение", "Юрист объясняет, что в лицензии AI-инструментов реально означает «коммерческое использование»",
     "Юридическая экспертиза как credential для разбора мелкого шрифта → примеры формулировок → практический вывод",
     "Редкая юридическая ниша резко повышает доверие и шанс сохранения",
     "Высокие сохранения как справочный материал, репосты юридическим отделам"),
]

SS = [
    ("Сравнение", "Сравнил ChatGPT и Claude в одном письме — разница в тоне шокировала. Сохрани — пригодится в понедельник",
     "Параллельная задача для обоих инструментов → явная разница в результате → CTA на сохранение под конкретный повод",
     "Прямой CTA с конкретным поводом сохранения резко повышает saves и sends",
     "Высокие сейвы и репосты в директ перед началом рабочей недели"),
    ("Туториал", "Перешли тому, кто до сих пор печатает письма с нуля",
     "Демонстрация AI-шаблона письма → явный адресат пересылки в подписи → конкретная экономия времени",
     "Чёткий адресат снижает трение для решения поделиться",
     "Высокий sends per reach, комментарии с тегами коллег"),
    ("Продуктивность", "3 AI-инструмента, которые твоя команда ещё не знает",
     "Listicle-формат с конкретными названиями → демонстрация каждого за 5–7 секунд → итоговая карточка со списком",
     "Listicle-формат изначально оптимизирован под пересылку команде или начальнику",
     "Высокие сейвы, репосты в рабочие чаты целиком"),
    ("Шоукейс", "Это нужно показать дизайнеру в твоей команде прямо сейчас",
     "Wow-результат AI-генерации → явный адресат «дизайнеру в команде» → CTA переслать",
     "Узкий целевой адресат повышает релевантность пересылки",
     "Высокие репосты в директ конкретным профессиям"),
    ("Сравнение", "Какой AI-инструмент выбрала бы твоя команда — голосуй в комментариях",
     "Сравнение двух вариантов без явного вывода → призыв голосовать → ответ автора в комментариях позже",
     "Открытый вопрос провоцирует комментарии и обсуждение в обоих лагерях",
     "Высокая комментарийная активность, рост охвата через дискуссию"),
    ("Юмор", "Отправь это тому, кто всё ещё спорит, что AI бесполезен",
     "Юмористический контраст ожидание/реальность → явный адресат-скептик → ирония в payoff",
     "Провокационный адресат создаёт повод для дружеского спора в личке",
     "Высокие репосты как повод для шутливого спора с конкретным человеком"),
    ("Продуктивность", "Сохрани это до следующего дедлайна — пригодится",
     "Конкретный AI-приём под цейтнот → явная привязка к будущей ситуации → CTA на сохранение",
     "Привязка к будущему событию создаёт надёжный отложенный повод вернуться",
     "Высокие сейвы с долгим временем удержания (return visits)"),
    ("AI для бизнеса", "Покажи это своему руководителю — это сэкономит бюджет команды",
     "Конкретный кейс экономии → явный адресат (руководитель) → измеримая выгода в payoff",
     "Прямая финансовая выгода для адресата снижает трение для пересылки наверх",
     "Высокие репосты внутри корпоративных чатов, сохранения для встреч"),
    ("Сравнение", "5 AI-инструментов одной картинкой — сохрани как шпаргалку",
     "Инфографика-сводка в одном кадре → краткое объяснение каждого пункта → явный CTA «сохрани как шпаргалку»",
     "Формат «шпаргалка» специально оптимизирован под высокую долю сохранений",
     "Очень высокие сейвы, относительно низкие лайки — характерный паттерн референсного контента"),
    ("Туториал", "Скинь другу-фрилансеру — это изменит его расчёт времени",
     "AI-приём для тайм-трекинга → явный адресат-фрилансер → конкретная цифра сэкономленных часов",
     "Узкая профессиональная аудитория с чётким адресатом повышает релевантность пересылки",
     "Высокие репосты во фриланс-сообщества"),
    ("AI-новости", "Перешли тому, кто пропустил этот релиз — он важный",
     "Сжатая новость с конкретной деталью → явный адресат «кто пропустил» → CTA переслать",
     "Роль «быть первым, кто сообщил новость другу» — сильный мотиватор пересылки",
     "Высокие репосты сразу после публикации, пик в первые 2 часа"),
    ("Шоукейс", "Тегни того, кому нужно это увидеть для вдохновения",
     "Впечатляющий AI-результат → явный призыв тегнуть → открытый комментарий для тегов",
     "Прямой призыв к тегу создаёт публичный, видимый sharing-сигнал",
     "Высокая комментарийная активность с тегами, вторичный охват через тегнутых"),
    ("Продуктивность", "Это для группового чата с командой — отправь не глядя",
     "Универсальный AI-лайфхак, подходящий любой команде → призыв «отправь не глядя» → минимум объяснений",
     "Сниженный порог решения увеличивает импульсивные репосты",
     "Высокий sends per reach в первые часы после публикации"),
    ("Сравнение", "Спор в комментариях: какой AI отвечает честнее на сложные вопросы",
     "Провокационное сравнение ответов на спорный вопрос → приглашение к дискуссии → автор не даёт финальный вердикт",
     "Намеренно открытый спор без вердикта продлевает обсуждение и органический охват",
     "Очень высокая комментарийная активность, длинные треды"),
    ("Юмор", "Если ты понял этот мем про AI — ты в теме",
     "Нишевый юмор, понятный только активным пользователям AI-инструментов → визуальная шутка → CTA тегнуть «своих»",
     "Инсайдерский юмор создаёт эффект эксклюзивного клуба, который хочется показать «своим»",
     "Высокие репосты внутри нишевых комьюнити-чатов"),
    ("AI для бизнеса", "Отправь в чат основателей — редкий честный разбор AI-инструментов для бизнеса",
     "Развёрнутый честный разбор без рекламной подачи → явный адресат (чат основателей) → конкретные выводы",
     "Адресность к узкой профессиональной группе повышает релевантность шеринга",
     "Высокие репосты в закрытые бизнес-чаты, сохранения как референс"),
    ("Промо", "Если знаешь того, кому пора сменить профессию — это видео для него",
     "История трансформации через продукт → явный адресат «тому, кому пора» → CTA переслать с приглашением",
     "Эмоционально значимый адресат — сильный повод для личной, осмысленной пересылки",
     "Высокие конверсионные переходы по ссылке от пересланных контактов"),
]

ST = [
    ("История", "Чуть не сорвал дедлайн — нашёл это в 11 вечера",
     "Hook: тревожная завязка → Problem: отчёт на 6 часов работы → Story: поиск решения ночью → Resolution: закрыл за 12 минут, показано как именно",
     "Завершённая личная история с конкретным числом вызывает эмпатию и желание поделиться «со мной было так же»",
     "Высокий completion rate, эмоциональные комментарии с похожими историями"),
    ("Туториал", "Меня чуть не уволили за опоздавший отчёт — а потом я нашёл способ его не бояться",
     "Hook: личный провал → Problem: повторяющаяся проблема со временем → Story: неудачные попытки решений → Resolution: финальный AI-приём и итог",
     "Уязвимая личная история создаёт сильную эмоциональную связь со зрителем",
     "Высокие репосты с личными комментариями о похожем опыте"),
    ("AI для бизнеса", "Мы чуть не потеряли клиента из-за медленного ответа — вот что изменило всё",
     "Hook: бизнес-кризис → Problem: медленная коммуникация с клиентом → Story: внедрение AI-ассистента под давлением сроков → Resolution: клиент остался, цифры ответа",
     "Бизнес-история с реальной ставкой резонирует с предпринимательской аудиторией",
     "Высокие сохранения как кейс для собственной команды"),
    ("Шоукейс", "Я не верил, что AI нарисует то, что у меня в голове — пока не увидел это",
     "Hook: скепсис как завязка → Problem: трудно объяснить идею художнику → Story: промптинг с правками → Resolution: финальное изображение, совпавшее с задумкой",
     "Дуга «от скепсиса к удивлению» вызывает желание показать другим доказательство",
     "Высокий completion rate, комментарии с собственными попытками повторить"),
    ("Продуктивность", "Я работал по выходным три месяца подряд — потом изменил один процесс",
     "Hook: усталость как завязка → Problem: переработки выходного дня → Story: AI-автоматизация рутинных задач → Resolution: первые свободные выходные, конкретный эффект",
     "История про выгорание и восстановление баланса находит широкий эмоциональный отклик",
     "Высокие репосты близким с личными комментариями"),
    ("Юмор", "Доверил AI ответить клиенту вместо меня — вот что произошло",
     "Hook: рискованное решение → Problem: усталость отвечать однотипно → Story: AI отвечает, начинаются неожиданности → Resolution: комедийная развязка с выводом",
     "Комедийная история с риском и неожиданной развязкой — формат для пересылки «ты не поверишь»",
     "Высокие репосты с подписью «вот это поворот»"),
    ("Туториал", "Полгода училась монтажу руками — а потом узнала про эту функцию",
     "Hook: долгий путь вручную как завязка → Problem: время на ручной монтаж → Story: случайное открытие AI-функции → Resolution: сравнение времени до/после",
     "История «зря потраченного времени» резонирует с теми, кто сейчас делает так же",
     "Высокие сохранения, комментарии «а я только что узнал(а)»"),
    ("AI-новости", "Я был скептиком насчёт этого обновления — пока не протестировал его на реальном проекте",
     "Hook: публичный скепсис как завязка → Problem: ожидание разочарования от хайпа → Story: тест на реальной задаче → Resolution: честный вывод, изменивший мнение",
     "Дуга «от скепсиса к признанию» воспринимается честнее безусловного хайпа, повышая доверие к шерингу",
     "Высокие сохранения, доверительные комментарии"),
    ("Шоукейс", "Клиент попросил невозможное — у меня было 2 часа на это",
     "Hook: давление времени как завязка → Problem: невыполнимый на первый взгляд запрос → Story: AI-инструменты под дедлайном → Resolution: сданная работа, реакция клиента",
     "История «под давлением» с конкретными ставками держит внимание через нарастающее напряжение",
     "Высокий completion rate, комментарии с похожими рабочими историями"),
    ("Продуктивность", "У меня было 47 непрочитанных писем каждое утро — пока я не сделал одно",
     "Hook: конкретная цифра боли как завязка → Problem: накопление и стресс от почты → Story: настройка AI-сортировки писем → Resolution: новое утро с конкретным числом писем",
     "Конкретные узнаваемые цифры делают историю максимально относимой к опыту зрителя",
     "Высокие репосты с подписью «это про меня»"),
    ("AI для бизнеса", "Стартап на грани закрытия — один процесс на AI всё изменил",
     "Hook: драматичная завязка про выживание бизнеса → Problem: нехватка ресурсов на ручную работу → Story: автоматизация ключевого процесса → Resolution: бизнес выжил, цифры роста",
     "История выживания бизнеса с высокими ставками вызывает сильный эмоциональный отклик у предпринимателей",
     "Высокие сохранения, репосты в стартап-сообщества"),
    ("Туториал", "Меня высмеяли за вопрос на созвоне — теперь я готовлюсь иначе",
     "Hook: неловкая личная ситуация как завязка → Problem: страх задать «глупый» вопрос неподготовленным → Story: подготовка к встречам с AI-инструментом → Resolution: уверенность на следующем созвоне",
     "Уязвимая социальная история вызывает сильную эмпатию и желание помочь другим избежать этого",
     "Высокие репосты коллегам перед важными встречами"),
    ("Сравнение", "Я потратил $200 на неправильный инструмент, прежде чем нашёл этот",
     "Hook: финансовая ошибка как завязка → Problem: разочарование в первой покупке → Story: поиск альтернативы методом проб → Resolution: более дешёвое и эффективное решение",
     "История о финансовой ошибке и её исправлении экономит деньги зрителю — сильный повод поделиться",
     "Высокие сохранения, комментарии с собственным опытом трат"),
    ("Юмор", "Попросил AI написать письмо боссу — отправил не читая",
     "Hook: рискованное решение как завязка → Problem: лень/спешка → Story: что было в письме на самом деле → Resolution: комедийная развязка с выводом «всегда проверяй»",
     "Комедийная история про реальный риск с относимой ошибкой — формат для пересылки «у меня чуть так не было»",
     "Высокие репосты с шутливыми комментариями о похожих рисках"),
    ("Продуктивность", "Я говорила «у меня нет на это времени» полгода — потом нашла 2 часа в день",
     "Hook: повторяющаяся отговорка как завязка → Problem: ощущение нехватки времени → Story: AI-аудит собственного расписания → Resolution: конкретные найденные 2 часа",
     "История самопознания через данные резонирует с широкой аудиторией перегруженных людей",
     "Высокие сохранения, репосты с личными признаниями в комментариях"),
    ("Промо", "Год назад я не знала, что такое промпт — сегодня это моя профессия",
     "Hook: контраст «было / стало» как завязка → Problem: страх начать с нуля → Story: путь обучения через продукт → Resolution: текущий профессиональный результат",
     "Трансформационная история «с нуля до результата» — мощный социальный proof для тех, кто сомневается",
     "Высокие переходы по ссылке от пересланных историй, комментарии «как начать»"),
]

BANK = ([("ME",) + t for t in ME] + [("CT",) + t for t in CT] + [("IS",) + t for t in IS] +
         [("AC",) + t for t in AC] + [("SS",) + t for t in SS] + [("ST",) + t for t in ST])
assert len(BANK) == 100, f"Content Bank must have 100 rows, has {len(BANK)}"

ws9 = wb.create_sheet("📚 09 Content Bank")
ws9.sheet_view.showGridLines = False
ws9.sheet_properties.tabColor = L_ENG
wid(ws9, [4, 16, 34, 16, 38, 32, 28])

r = header_block(
    ws9, 1, 7,
    "LAYER 3 · CONTENT ENGINE — DELIVERABLE",
    "📚  SHORT-FORM CONTENT BANK — 100 СЦЕНАРИЕВ",
    "Ниша: AI / продуктивность / цифровые инструменты. Каждый сценарий закодирован "
    "одним из 6 принципов из 03 Psychology — это не случайные идеи, а выход модели.",
    L_ENG,
)

legend = "  ".join(f"■ {name}" for name, _, _ in PRINC_META.values())
r = note(ws9, r, 1, 7, "Легенда принципов:  " + legend, h=18, bg=WHITE, fc=MUTED)

r = section(ws9, r, 1, 7, "  100 СЦЕНАРИЕВ", L_ENG)
heads = ["#", "Тип контента", "Hook (0–2 сек)", "Принцип", "Структура H→P→S→R", "Почему завирусится", "Ожидаемое поведение зрителя"]
for ci, h in enumerate(heads, 1):
    hd(ws9, r, ci, h)
rh(ws9, r, 26)
r += 1
bank_start = r
for i, (pkey, pillar, hook, structure, virality, engagement) in enumerate(BANK):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    pname, pbg, pfc = PRINC_META[pkey]
    cl(ws9, rr, 1, i + 1, bold=True, bg=bg, align="center", fc=ACCENT, sz=8.5)
    cl(ws9, rr, 2, pillar, bg=bg, fc=TEXT, sz=8.3, bold=True)
    cl(ws9, rr, 3, hook, bg=bg, fc=INK, sz=8.5, bold=True)
    cl(ws9, rr, 4, pname, bg=pbg, fc=pfc, sz=8.3, bold=True, align="center")
    cl(ws9, rr, 5, structure, bg=bg, fc=TEXT, sz=8)
    cl(ws9, rr, 6, virality, bg=bg, fc=TEXT, sz=8)
    cl(ws9, rr, 7, engagement, bg=bg, fc=MUTED, sz=8, italic=True)
    rh(ws9, rr, 58)
bank_end = r + len(BANK) - 1

freeze_below_header(ws9, bank_start)
print(f"Sheet 9 (Content Bank) built — {len(BANK)} scripts, rows {bank_start}-{bank_end}")

# ════════════════════════════════════════════════════════════════════
# SHEET 10 — REALITY CHECK (appendix)
# ════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("🔬 10 Reality Check")
ws10.sheet_view.showGridLines = False
ws10.sheet_properties.tabColor = L_REAL
wid(ws10, [34, 46, 30])

r = header_block(
    ws10, 1, 3,
    "ПРИЛОЖЕНИЕ · REALITY CHECK",
    "🔬  ПОЧЕМУ AI-КОНТЕНТ МОЖЕТ ТЕРЯТЬ ОХВАТ В INSTAGRAM",
    "Система производит объём — но объём без понимания ранжирования Instagram "
    "превращается в спам-паттерн, который алгоритм подавляет. Раздел основан "
    "на актуальных данных об алгоритме Reels 2026.",
    L_REAL,
)

r = section(ws10, r, 1, 3, "  А. ФАКТОРЫ, КОТОРЫЕ МОГУТ ЗАНИЖАТЬ ОХВАТ AI-КОНТЕНТА", L_REAL)
hd(ws10, r, 1, "Фактор"); hd(ws10, r, 2, "Почему это риск для этой системы"); hd(ws10, r, 3, "Уровень риска")
rh(ws10, r, 22)
r += 1
RISK = {"Высокий": (BAD_BG, BAD), "Средний": (WARN_BG, WARN), "Низкий": (GOOD_BG, GOOD)}
risks = [
    ("Классификатор оригинальности (Originality Score, с окт. 2025)",
     "Шаблонный AI-контент с одинаковыми визуальными паттернами и войсовером на 15 аккаунтах "
     "одновременно статистически похож на репост/aggregator-паттерн. Аккаунты с 10+ репостами "
     "за 30 дней полностью исключаются из рекомендаций.", "Высокий"),
    ("Скриншоты интерфейсов без трансформации",
     "Прямой скринкаст чужого приложения без монтажной обработки и личного комментария может "
     "читаться алгоритмом как контент «не с этой платформы», а не как оригинальная съёмка.", "Средний"),
    ("AI-info / AI-creator лейблы на органике",
     "Instagram официально заявляет, что лейбл AI-info на отдельном посте не влияет на "
     "дистрибуцию органического контента. Риск реален в первую очередь для рекламного/"
     "спонсорского контента с реалистичным AI-голосом или сценами, оценёнными как вводящие "
     "в заблуждение — там зафиксированы падения охвата до ~80%.", "Средний"),
    ("Однородный синтетический «акустический отпечаток»",
     "Полностью синтетический войсовер на всех 100 видео в день создаёт одинаковое звучание "
     "по всей экосистеме. Это не штрафуется алгоритмом напрямую, но снижает ощущение «живого "
     "автора» у зрителя и может снижать completion rate, который Instagram использует как "
     "ключевой сигнал ранжирования.", "Средний"),
    ("Объём публикаций без пропорционального роста engagement",
     "Алгоритм сравнивает baseline-engagement аккаунта с историей. Постинг ради количества без "
     "роста вовлечённости размывает средний engagement и наказывается понижением охвата — "
     "«два сильных поста в неделю лучше семи посредственных».", "Высокий"),
    ("EU AI Act, Статья 50 (обязательна с 2 авг. 2026)",
     "В ЕС AI-контент без существенной человеческой редактуры подлежит обязательной маркировке. "
     "Контент, прошедший человеческое редактирование/ответственность, освобождён от требования "
     "— это прямой аргумент в пользу QC-шлюза на этапе 7 пайплайна (см. 05 Pipeline).", "Низкий"),
]
for i, (factor, why, risk) in enumerate(risks):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws10, rr, 1, factor, bold=True, bg=bg, fc=L_REAL, sz=9)
    cl(ws10, rr, 2, why, bg=bg, fc=TEXT, sz=8.5)
    rbg, rfc = RISK[risk]
    cl(ws10, rr, 3, risk, bg=rbg, fc=rfc, bold=True, align="center")
    rh(ws10, rr, 64)
r += len(risks) + 1

r = section(ws10, r, 1, 3, "  Б. ЧТО РЕАЛЬНО ВЛИЯЕТ НА ОХВАТ В 2026", SLATE)
hd(ws10, r, 1, "Сигнал ранжирования"); hd(ws10, r, 2, "Что это значит на практике"); hd(ws10, r, 3, "Вес")
rh(ws10, r, 22)
r += 1
WEIGHT = {"Критический": (ACCENT_BG, ACCENT), "Высокий": (GOOD_BG, GOOD), "Средний": (WARN_BG, WARN)}
real_factors = [
    ("Watch time / completion rate", "Подтверждено Instagram как фактор №1. Короткие, плотные ролики с удержанием первых 3 секунд важнее количества лайков.", "Критический"),
    ("Sends per reach (репосты в директ)", "Вес в 3–5 раз выше обычных лайков для охвата за пределами подписчиков — главный множитель холодного охвата.", "Критический"),
    ("Saves (сохранения)", "Сигнал долгосрочной ценности, особенно для образовательного и референсного контента — то есть для туториалов.", "Высокий"),
    ("Глубина комментариев", "Длина и осмысленность диалога важнее количества; ответ в первые 30 минут — отдельный сигнал ранжирования.", "Высокий"),
    ("Originality (не репост/не watermark)", "Оригинальный контент получает на 40–60% больше дистрибуции, чем репосты; watermark с других платформ детектируется и понижается.", "Критический"),
    ("«Your Algorithm» topic controls (с дек. 2025)", "Пользователи явно выбирают темы в настройках. Контент без чёткой нишевой принадлежности становится «невидимым» — прямое обоснование нишевой сегментации в 02 Accounts.", "Высокий"),
    ("Account-level baseline engagement", "Алгоритм сравнивает каждый новый пост со средним engagement аккаунта — устойчиво слабые посты тянут вниз весь аккаунт.", "Средний"),
]
for i, (sig, mean, w) in enumerate(real_factors):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws10, rr, 1, sig, bold=True, bg=bg, fc=TEXT, sz=9)
    cl(ws10, rr, 2, mean, bg=bg, fc=TEXT, sz=8.5)
    wbg, wfc = WEIGHT[w]
    cl(ws10, rr, 3, w, bg=wbg, fc=wfc, bold=True, align="center")
    rh(ws10, rr, 46)
r += len(real_factors) + 1

r = section(ws10, r, 1, 3, "  В. КАК ЭТО КОМПЕНСИРОВАТЬ В СТРАТЕГИИ", L_REAL)
hd(ws10, r, 1, "Действие"); hd(ws10, r, 2, "Почему это работает"); hd(ws10, r, 3, "Где в системе")
rh(ws10, r, 22)
r += 1
compensate = [
    ("Трансформировать, а не просто записывать", "Монтаж + личный комментарий + субтитры поверх любого скринкаста переводит контент из категории «репост» в «оригинальный».", "05 Pipeline, шаг 6"),
    ("Живой голос на части аккаунтов", "Снижает однородность «акустического отпечатка», повышает ощущение creator presence хотя бы на флагманских каналах.", "06 Team, 07 AI Stack"),
    ("Тестировать форматы на холодную аудиторию перед широким постингом", "Снижает риск размытия baseline-engagement аккаунта слабым форматом до его публикации основной аудитории.", "01 Strategy, принцип 3"),
    ("Мониторить non-follower reach % ежедневно", "Канареечная метрика: падение сигналит о проблеме раньше, чем падают абсолютные просмотры.", "08 KPI Dashboard"),
    ("Держать 1 чёткую нишу на аккаунт", "Прямой ответ на topic-controls — контент должен однозначно попадать в одну выбираемую пользователем категорию.", "02 Accounts"),
    ("Вариативность подачи между 15 аккаунтами", "Разные хуки/темп/монтаж снижают вероятность паттерн-детекции «шаблонности» по экосистеме целиком.", "04 Content Engine"),
    ("CTA на вопрос + ответ в первые 30 минут", "Прямой, подтверждённый алгоритмический сигнал — глубина комментариев весит больше их количества.", "06 Team — Менеджер аккаунтов"),
    ("QC-шлюз = «существенная человеческая редактура»", "Снимает риск обязательной AI-маркировки по EU AI Act Статья 50 и одновременно поднимает качество.", "05 Pipeline, шаг 7"),
]
for i, (action, why, where) in enumerate(compensate):
    rr = r + i
    bg = ROW_ALT if i % 2 == 0 else WHITE
    cl(ws10, rr, 1, action, bold=True, bg=bg, fc=L_REAL, sz=8.8)
    cl(ws10, rr, 2, why, bg=bg, fc=TEXT, sz=8.5)
    cl(ws10, rr, 3, where, bg=bg, fc=MUTED, sz=8.3, italic=True, align="center")
    rh(ws10, rr, 40)
r += len(compensate) + 1

r = note(ws10, r, 1, 3,
         "Источники: Adam Mosseri (Instagram Head), публичные заявления янв. 2025 – июнь 2026; "
         "обзоры алгоритма Reels 2026 от Hootsuite, Buffer, Later, Sprout Social; обзоры политики "
         "маркировки AI-контента Meta и EU AI Act, 2026. Цифры и пороги могут измениться — "
         "перепроверяйте раз в квартал.", h=30)

freeze_below_header(ws10, 5)
print("Sheet 10 (Reality Check) built")

# ════════════════════════════════════════════════════════════════════
# SAVE — portable path, runs unmodified on macOS / Linux / Windows
# ════════════════════════════════════════════════════════════════════
for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")







