"""ExcelStorage — конкретная реализация хранилища на openpyxl (Infrastructure layer).

Раздел 10 архитектуры: реализация StorageInterface.
Открывает существующую книгу, находит/создаёт лист Leads,
дописывает строку, сохраняет файл.

Раздел 14: load_workbook → modify → save внутри критической секции asyncio.Lock.
"""

from __future__ import annotations

import asyncio
from concurrent.futures import ThreadPoolExecutor
from typing import List, Optional

from loguru import logger
from openpyxl import Workbook, load_workbook

from bot.models.lead import Lead
from bot.storage.base import SaveResult, StorageInterface
from bot.storage.exceptions import StorageUnavailableError
from bot.storage.lock import AsyncIOLock

# Заголовки колонок листа Leads
_LEADS_HEADERS = [
    "telegram_id",
    "telegram_username",
    "email",
    "submitted_at",
    "source",
]

# Для run_in_executor — отдельный пул потоков для I/O
_IO_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="xlsx_io")


class ExcelStorage(StorageInterface):
    """Реализация StorageInterface на openpyxl.

    Работает по схеме: загрузить весь файл → изменить в памяти → сохранить весь файл.
    Все операции сериализуются через AsyncIOLock.
    """

    def __init__(self, file_path: str, sheet_name: str = "Leads") -> None:
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._lock = AsyncIOLock()
        self._executor = _IO_EXECUTOR

    def _run_sync(self, func, *args):
        """Выполнить синхронную openpyxl-операцию в thread pool."""
        loop = asyncio.get_running_loop()
        return loop.run_in_executor(self._executor, func, *args)

    @staticmethod
    def _load_workbook_sync(file_path: str) -> Workbook:
        """Синхронная загрузка workbook. Вызывается в executor.

        Если файла нет — создаётся новый workbook (раздел 4.5 архитектуры:
        бот создаёт лист Leads при первом сохранённом лиде, если файла нет).
        """
        import os
        if not os.path.exists(file_path):
            # Убедиться, что родительский каталог существует
            parent = os.path.dirname(file_path)
            if parent:
                os.makedirs(parent, exist_ok=True)
            logger.info(f"Excel file not found, creating new: {file_path}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"

            ws.append([
                "telegram_id",
                "telegram_username",
                "email",
                "submitted_at",
                "source",
            ])

            wb.save(file_path)

            return wb
        try:
            return load_workbook(file_path)
        except Exception as exc:
            logger.error(f"Failed to load workbook {file_path}: {exc}")
            raise StorageUnavailableError(f"Не удалось открыть файл: {exc}") from exc

    @staticmethod
    def _save_workbook_sync(wb: Workbook, file_path: str) -> None:
        """Синхронное сохранение workbook. Вызывается в executor."""
        try:
            wb.save(file_path)
        except PermissionError as exc:
            logger.error(f"Permission denied saving {file_path}: {exc}")
            raise StorageUnavailableError(f"Нет прав на запись в файл: {exc}") from exc
        except Exception as exc:
            logger.error(f"Failed to save workbook {file_path}: {exc}")
            raise StorageUnavailableError(f"Ошибка сохранения: {exc}") from exc

    def _ensure_leads_sheet(self, wb: Workbook) -> None:
        """Найти или создать лист Leads. Добавить заголовки при первом создании."""
        if self._sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(self._sheet_name)
            ws.append(_LEADS_HEADERS)
            logger.info(f"Created sheet '{self._sheet_name}' with headers")

    def _lead_to_row(self, lead: Lead) -> list:
        """Преобразовать Lead в строку для ws.append()."""
        return [
            lead.telegram_id,
            lead.telegram_username or "",
            lead.email,
            lead.submitted_at.isoformat(),
            lead.source,
        ]

    def _read_emails_sync(self, wb: Workbook, sheet_name: str) -> set:
        """Синхронное чтение email из листа. Вызывается в executor."""
        if sheet_name not in wb.sheetnames:
            return set()
        ws = wb[sheet_name]
        emails = set()
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
            if row[0] is not None:
                emails.add(str(row[0]).strip().lower())
        return emails

    def _count_rows_sync(self, wb: Workbook, sheet_name: str) -> int:
        """Синхронный подсчёт строк данных (без заголовка)."""
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        return max(0, ws.max_row - 1)

    def _read_leads_sync(self, wb: Workbook, sheet_name: str, limit: int, offset: int) -> list:
        """Синхронное чтение строк с заданным offset и limit."""
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        leads = []
        for row in ws.iter_rows(min_row=offset + 2, max_col=5, values_only=True):
            if len(leads) >= limit:
                break
            if row[0] is None:
                continue
            from datetime import datetime, timezone

            submitted_at = row[3]
            if isinstance(submitted_at, str):
                try:
                    submitted_at = datetime.fromisoformat(submitted_at)
                except ValueError:
                    submitted_at = datetime.now(timezone.utc)
            elif not isinstance(submitted_at, datetime):
                submitted_at = datetime.now(timezone.utc)

            leads.append(
                Lead(
                    telegram_id=int(row[0]),
                    telegram_username=row[1] if row[1] else None,
                    email=str(row[2]),
                    submitted_at=submitted_at,
                    source=str(row[4]) if row[4] else "bot_v1",
                )
            )
        return leads

    async def save_lead(self, lead: Lead) -> SaveResult:
        """Атомарно добавить одну запись в хранилище.

        Критическая секция: load_workbook → modify → save (раздел 14.3).
        """
        async with self._lock.acquire():
            try:
                wb = await self._run_sync(self._load_workbook_sync, self._file_path)
                self._ensure_leads_sheet(wb)
                ws = wb[self._sheet_name]
                ws.append(self._lead_to_row(lead))
                await self._run_sync(self._save_workbook_sync, wb, self._file_path)
                logger.info(f"Lead saved: {lead.email} (tg_id={lead.telegram_id})")
                return SaveResult(success=True)
            except StorageUnavailableError:
                raise
            except Exception as exc:
                logger.error(f"Unexpected error saving lead: {exc}")
                raise StorageUnavailableError(f"Ошибка при сохранении: {exc}") from exc

    async def email_exists(self, email: str) -> bool:
        """Проверить уникальность email.

        Входит в ту же критическую секцию, что и save_lead (D004).
        Это устраняет TOCTOU-гонку.
        """
        async with self._lock.acquire():
            try:
                wb = await self._run_sync(self._load_workbook_sync, self._file_path)
                emails = await self._run_sync(self._read_emails_sync, wb, self._sheet_name)
                return email.lower() in emails
            except StorageUnavailableError:
                raise
            except Exception as exc:
                logger.error(f"Unexpected error checking email existence: {exc}")
                raise StorageUnavailableError(f"Ошибка при проверке email: {exc}") from exc

    async def count_leads(self) -> int:
        """Текущее количество сохранённых записей."""
        async with self._lock.acquire():
            try:
                wb = await self._run_sync(self._load_workbook_sync, self._file_path)
                return await self._run_sync(self._count_rows_sync, wb, self._sheet_name)
            except StorageUnavailableError:
                return 0
            except Exception as exc:
                logger.error(f"Error counting leads: {exc}")
                return 0

    async def list_leads(self, limit: int = 50, offset: int = 0) -> List[Lead]:
        """Постраничная выборка лидов."""
        async with self._lock.acquire():
            try:
                wb = await self._run_sync(self._load_workbook_sync, self._file_path)
                return await self._run_sync(
                    self._read_leads_sync, wb, self._sheet_name, limit, offset
                )
            except StorageUnavailableError:
                return []
            except Exception as exc:
                logger.error(f"Error listing leads: {exc}")
                return []

    async def health_check(self) -> bool:
        """Проверка доступности хранилища."""
        try:
            wb = await self._run_sync(self._load_workbook_sync, self._file_path)
            # Успешная загрузка = хранилище доступно
            return True
        except StorageUnavailableError:
            return False
        except Exception:
            return False
